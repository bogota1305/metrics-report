import os
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from io import BytesIO

from modules.colors import lighten_color
from uploadCloud import upload_to_drive, upload_to_dropbox

def save_dataframe_to_excel(output_dir, output_file, data, sheet_name, columns_to_plot, colors, grafico_positions, dropbox_var=False, drive_var=False):
   
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Construir la ruta completa del archivo
    full_path = os.path.join(output_dir, f"{output_file}.xlsx")

    wb = Workbook()
    wbLineChart = line_chart(wb, sheet_name, data, columns_to_plot, colors, grafico_positions)
    wbLineChart.save(full_path)
    urls = []
    url = ''

    if(output_file != 'Payment Errors' and dropbox_var):
        url = upload_to_dropbox(full_path, dropbox_path=f"/MyReports/{output_dir}/{output_file}.xlsx") 
    if(output_file != 'Payment Errors' and drive_var):
        url = upload_to_drive(full_path, folder_id="1F1VZxlp5IxkQEo4WD0Bt8VEJZ28OhGut")
    
    urls.insert(0, url)

    return urls

    

    # Hoja 1: Resumen General

def line_chart(wb, sheet_name, data, columns_to_plot, colors, grafico_positions):
    ws1 = wb.active
    ws1.title = sheet_name
    for r in dataframe_to_rows(data, index=False, header=True):
        ws1.append(r)

    # Colorear la fila "Total"
    for cell in ws1[ws1.max_row]:  # Iterar sobre todas las celdas de la última fila
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    filtered_summary = data[data['date'] != 'Total']

    for col_idx, (column, color, position) in enumerate(zip(columns_to_plot, colors, grafico_positions)):
        plt.figure(figsize=(10, 6))
        plt.plot(filtered_summary['date'], filtered_summary[column], marker='o', linestyle='-', color=color)
        
        promedio = filtered_summary[column].mean()
        plt.axhline(promedio, color="gray", linestyle="--", label=f"Promedio: {promedio:.2f}")

        plt.title(f"{column} by Day")
        plt.xlabel("Day")
        plt.ylabel(column.replace('_', ' ').title())
        plt.legend()
        plt.xticks(rotation=45)

        chart_image_path = f"{column}_chart.png"
        plt.tight_layout()
        plt.savefig(chart_image_path)
        plt.close()

        img = Image(chart_image_path)
        img.width = 600
        img.height = 400
        ws1.add_image(img, position)

        light_color = lighten_color(color, factor=0.5)
        column_letter = chr(65 + col_idx + 1)
        title_cell = ws1[f"{column_letter}1"]
        title_cell.fill = PatternFill(start_color=light_color[1:], end_color=light_color[1:], fill_type="solid")

    # Ajustar tamaño de columnas en la hoja 1
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws1.column_dimensions[column].width = max(max_length + 2, 15)

    return wb

def save_error_reasons_with_chart(output_dir, file_name, error_reasons, is_payment, dropbox_var, drive_var):
    """
    Guarda las razones de error en una nueva hoja de Excel, pinta las celdas
    con colores dinámicos según la razón de error, añade un gráfico de barras
    creado con matplotlib y ajusta el ancho de las columnas automáticamente.
    """
    from openpyxl import load_workbook

    reason_type = 'reason'
    count = 'cancelation_count'
    xLabel = 'Cancelation Type'
    yLabel = 'Cancelation Count'
    sheetName = 'Cancelation Reasons'
    title = 'Cancelation Reasons'

    if(is_payment):
        reason_type = 'decline_code'
        count = 'error_count'
        xLabel = 'Error Type'
        yLabel = 'Error Count'
        sheetName = 'Error Reasons'
        title = 'Payment Error Reasons'

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Construir la ruta completa del archivo
    full_path = os.path.join(output_dir, f"{file_name}.xlsx")

    # Generar colores únicos para cada tipo de error dinámicamente
    unique_reasons = error_reasons[reason_type].unique()
    colors = plt.cm.get_cmap('tab20', len(unique_reasons))  # Usamos 'tab10' para hasta 10 colores únicos
    color_map = {reason: tuple(int(c * 255) for c in colors(i)[:3]) for i, reason in enumerate(unique_reasons)}
    hex_color_map = {reason: f"{c[0]:02X}{c[1]:02X}{c[2]:02X}" for reason, c in color_map.items()}  # Colores en HEX

    # Crear un gráfico de barras con matplotlib
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(
        error_reasons[reason_type],
        error_reasons[count],
        color=[f"#{hex_color_map[reason]}" for reason in error_reasons[reason_type]]
    )
    ax.set_title(title, fontsize=14)
    ax.set_xlabel(xLabel)
    ax.set_ylabel(yLabel)
    plt.xticks(rotation=90)
    plt.tight_layout()

    # Guardar la imagen del gráfico en memoria
    img_data = BytesIO()
    plt.savefig(img_data, format='png')
    plt.close(fig)
    img_data.seek(0)

    # Abrir el archivo Excel existente
    workbook = load_workbook(full_path)

    # Crear una nueva hoja para las razones de error
    sheet_name = sheetName
    sheet = workbook.create_sheet(sheet_name)

    # Escribir los datos del DataFrame en la hoja
    for row_idx, row in enumerate(dataframe_to_rows(error_reasons, index=False, header=True), start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            # Pintar las celdas de la columna de 'decline_code' (excepto el header)
            if row_idx > 1 and col_idx == 1:  # Columna 1 = 'decline_code'
                reason = value
                if reason in hex_color_map:
                    cell.fill = PatternFill(
                        start_color=hex_color_map[reason],
                        end_color=hex_color_map[reason],
                        fill_type="solid"
                    )

    # Ajustar el ancho de las columnas según el contenido
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter  # Obtener la letra de la columna
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Margen adicional para un mejor ajuste
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Añadir el gráfico como imagen en la hoja
    img = Image(img_data)
    img.anchor = 'D5'  # Posición donde insertar el gráfico
    sheet.add_image(img)

    # Guardar los cambios en el archivo Excel
    workbook.save(full_path)
    urls = []
    url = ''
    
    if(dropbox_var):
        url = upload_to_dropbox(full_path, dropbox_path=f"/MyReports/{output_dir}/{file_name}.xlsx")
    if(drive_var):
        url = upload_to_drive(full_path, folder_id="1F1VZxlp5IxkQEo4WD0Bt8VEJZ28OhGut")

    urls.insert(0, url)
    
    return urls
    

def save_dataframe_to_excel_orders(output_dir, output_file, data, sheet_name, columns_to_plot, colors, grafico_positions, dropbox_var, drive_var):

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Construir la ruta completa del archivo
    full_path = os.path.join(output_dir, f"{output_file}.xlsx")

    # Crear el archivo Excel con gráficos
    wb = Workbook()
    wbLineChart = line_chart(wb, sheet_name, data, columns_to_plot, colors, grafico_positions)

    # Guardar el archivo en la carpeta especificada
    wbLineChart.save(full_path)

    url = ''
    
    if(dropbox_var):
        url = upload_to_dropbox(full_path, dropbox_path=f"/MyReports/{output_dir}/{output_file}.xlsx")
    if(drive_var):
        url = upload_to_drive(full_path, folder_id="1F1VZxlp5IxkQEo4WD0Bt8VEJZ28OhGut")
    
    return url


def save_dataframe_to_excel_ga4(percentages_table, percentages_previous_step, final_table_spaced_with_previous, nombre_salida, carpeta_salida, dropbox_var, drive_var):
    # Crear la carpeta si no existe
    if not os.path.exists(carpeta_salida):
        os.makedirs(carpeta_salida)

    # Definir las rutas completas para las gráficas y el archivo Excel
    chart_path = os.path.join(carpeta_salida, 'firstStep.png')
    chart_path_previous = os.path.join(carpeta_salida, 'stepPrevious.png')
    excel_path = os.path.join(carpeta_salida, nombre_salida)

    # Crear una gráfica a partir de la tabla de porcentajes, excluyendo la columna "Total"
    plt.figure(figsize=(10, 6))
    for step in percentages_table.index:
        short_label = step.replace(' (%)', '').split(' ')[0]
        numeric_values = (percentages_table.loc[step].drop('Total', errors='ignore').replace('%', '', regex=True).replace('', '0').astype(float))
        plt.plot(numeric_values, label=short_label)

    plt.title('Percentage Transition by Step')
    plt.xlabel('Days')
    plt.ylabel('Percentage')
    plt.legend(loc='upper left', bbox_to_anchor=(1, 1), fontsize='small')
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(chart_path)
    plt.close()

    # Crear una gráfica a partir de la tabla de porcentajes, excluyendo la columna "Total"
    plt.figure(figsize=(10, 6))
    for step in percentages_previous_step.index:
        short_label = step.split(' (')[0]
        numeric_values = percentages_previous_step.loc[step].drop('Total', errors='ignore').replace('%', '', regex=True).astype(float)
        plt.plot(numeric_values, label=short_label)

    plt.title('Percentage Step vs Previous')
    plt.xlabel('Days')
    plt.ylabel('Percentage')
    plt.legend(loc='upper left', bbox_to_anchor=(1, 1), fontsize='small')
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(chart_path_previous)
    plt.close()

    # Guardar la tabla final y las gráficas en un archivo Excel
    with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
        final_table_spaced_with_previous.to_excel(writer, sheet_name='Data', index=True, startrow=0, startcol=0)
        workbook = writer.book
        worksheet = writer.sheets['Data']

        # Ajustar el ancho de la columna de los pasos
        max_step_length = max(len(str(step)) for step in final_table_spaced_with_previous.index)
        worksheet.set_column(0, 0, max_step_length + 2)

        # Insertar las gráficas en el archivo Excel
        worksheet.insert_image('B18', chart_path)
        worksheet.insert_image('R18', chart_path_previous)
        
    # Eliminar las imágenes temporales
    try:
        os.remove(chart_path)
        os.remove(chart_path_previous)
    except OSError as e:
        print(f"Error al eliminar las imágenes: {e}")
    
    urls = []
    url = ''

    if(dropbox_var):
        url = upload_to_dropbox(excel_path, dropbox_path=f"/MyReports/{carpeta_salida}/{nombre_salida}") 
    if(drive_var):
        url = upload_to_drive(excel_path, folder_id="1F1VZxlp5IxkQEo4WD0Bt8VEJZ28OhGut")
        
    urls.insert(0, url)
    return urls

