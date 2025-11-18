import pandas as pd
import numpy as np
import json
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from modules.database_queries import execute_query

# Diccionario de shades (solo 30ml, excluyendo 45ml)
shades = {
    'IT00000000000000000000000000000021': '30ml Colorant - Light Auburn',
    'IT00000000000000000000000000000020': '30ml Colorant - Dark Auburn',
    'IT00000000000000000000000000000019': '30ml Colorant - Lightest Blond',
    'IT00000000000000000000000000000018': '30ml Colorant - Warm - Light Blond',
    'IT00000000000000000000000000000017': '30ml Colorant - Cool - Light Blond',
    'IT00000000000000000000000000000016': '30ml Colorant - Warm-Medium Blond',
    'IT00000000000000000000000000000015': '30ml Colorant - Cool-Medium Blond',
    'IT00000000000000000000000000000014': '30ml Colorant - Warm-Dark Blond',
    'IT00000000000000000000000000000013': '30ml Colorant - Dark Blond',
    'IT00000000000000000000000000000012': '30ml Colorant: Light Brown',
    'IT00000000000000000000000000000011': '30ml Colorant - Medium-Light Brown',
    'IT00000000000000000000000000000002': '30ml Colorant - Medium Brown',
    'IT00000000000000000000000000000010': '30ml Colorant - Medium-Dark Brown',
    'IT00000000000000000000000000000009': '30ml Colorant - Dark Brown',
    'IT00000000000000000000000000000008': '30ml Colorant - Soft-Black',
    'IT00000000000000000000000000000007': '30ml Colorant - Black',
    'IT00000000000000000000000000000006': '30ml Colorant - Jet-Black'
}

def extract_diagnostic_values(additional_fields):
    """
    Extrae los valores de diagnóstico (13, 14, 15) del campo additionalFields
    Devuelve una lista con los valores encontrados (solo el primero si hay múltiples)
    """
    if not additional_fields or pd.isna(additional_fields):
        return []
    
    try:
        diagnostic_data = json.loads(additional_fields)
        values_found = []
        
        # Buscar valores en diferentes formatos
        # Formato 1: Lista de diccionarios [{"value": 13, "variable": 4}, {"value": 14, "variable": 5}]
        if isinstance(diagnostic_data, list):
            for item in diagnostic_data:
                if isinstance(item, dict) and 'value' in item:
                    try:
                        value = int(item['value'])
                        if value in [13, 14, 15]:
                            values_found.append(value)
                    except (ValueError, TypeError):
                        continue
        
        # Formato 2: Diccionario con sub-diccionarios {"396654": {"value": "14", "variable": "4"}}
        elif isinstance(diagnostic_data, dict):
            for key, value_dict in diagnostic_data.items():
                # Si value_dict es un diccionario con clave 'value'
                if isinstance(value_dict, dict) and 'value' in value_dict:
                    try:
                        value = int(value_dict['value'])
                        if value in [13, 14, 15]:
                            values_found.append(value)
                    except (ValueError, TypeError):
                        continue
                # Si value_dict es una lista de diccionarios (caso que mencionas)
                elif isinstance(value_dict, list):
                    for item in value_dict:
                        if isinstance(item, dict) and 'value' in item:
                            try:
                                value = int(item['value'])
                                if value in [13, 14, 15]:
                                    values_found.append(value)
                            except (ValueError, TypeError):
                                continue
        
        # Tomar solo el primer valor si hay múltiples
        return values_found[:1] if values_found else []
        
    except json.JSONDecodeError:
        return []
    except Exception as e:
        print(f"Error procesando additionalFields: {e}")
        return []

def procesar_razon(razon):
    """
    Procesa la razón de cancelación para obtener más detalle en casos específicos
    Para 'I don't like my results' e 'I experienced a skin reaction' incluye hasta el segundo '->'
    Para otras razones, solo la parte antes del primer '->'
    """
    if pd.isna(razon):
        return "Sin razón especificada"
    
    razon_str = str(razon).strip()
    
    # Casos especiales que requieren más detalle
    special_cases = ["I don't like my results", "I experienced a skin reaction", "I don't like my results"]
    
    for case in special_cases:
        if razon_str.startswith(case):
            # Si contiene al menos dos '->', tomar hasta el segundo
            if razon_str.count('->') >= 2:
                parts = razon_str.split('->', 2)
                return f"{parts[0].strip()} -> {parts[1].strip()}"
            # Si solo tiene un '->', tomar todo
            elif '->' in razon_str:
                return razon_str
            # Si no tiene '->', devolver la razón completa
            else:
                return razon_str
    
    # Para otras razones, tomar solo la parte antes del primer '->'
    if '->' in razon_str:
        return razon_str.split('->')[0].strip()
    else:
        return razon_str

def agregar_porcentaje(valor):
    """Agrega el símbolo % a un valor numérico"""
    if pd.isna(valor):
        return ""
    return f"{valor}%"

def obtener_suscripciones_activas(startDate, endDate):
    """
    Obtiene las suscripciones activas por etnia y shade
    """
    item_ids = list(shades.keys())
    item_ids_str = "', '".join(item_ids)
    
    query = f"""
    SELECT  
        sub.id,
        sub.additionalFields->>"$.diagnostic" AS additionalFields,
        subIt.itemId
    FROM prod_sales_and_subscriptions.subscriptions sub
    JOIN bi.fact_orders fo ON sub.id = fo.subscription_id
    JOIN prod_sales_and_subscriptions.subscription_items subIt on sub.id = subIt.subscriptionId
    WHERE fo.status NOT IN ('CANCELLED','PAYMENT_ERROR')
    AND fo.created_at BETWEEN '{startDate}' AND '{endDate}'
    AND subIt.itemId IN ('{item_ids_str}')
    GROUP BY sub.id;
    """
    
    df_suscripciones = execute_query(query)
    
    # Procesar valores de diagnóstico para etnias
    df_suscripciones['diagnostic_values'] = df_suscripciones['additionalFields'].apply(extract_diagnostic_values)
    df_suscripciones['has_13'] = df_suscripciones['diagnostic_values'].apply(lambda x: 13 in x)
    df_suscripciones['has_14'] = df_suscripciones['diagnostic_values'].apply(lambda x: 14 in x)
    df_suscripciones['has_15'] = df_suscripciones['diagnostic_values'].apply(lambda x: 15 in x)
    
    # Crear columnas booleanas para cada shade
    for item_id, shade_name in shades.items():
        df_suscripciones[f'shade_{item_id[-4:]}'] = df_suscripciones['itemId'] == item_id
    
    return df_suscripciones

def crear_tabla_etnias(df_suscripciones):
    """
    Crea una tabla con el total de suscripciones por etnia
    """
    total_suscripciones = len(df_suscripciones)
    
    tabla_etnias = pd.DataFrame({
        'Etnia': ['CAUCASIAN', 'AFRICAN', 'ASIAN', 'Sin diagnóstico'],
        'Total Suscripciones': [
            df_suscripciones['has_13'].sum(),
            df_suscripciones['has_14'].sum(),
            df_suscripciones['has_15'].sum(),
            total_suscripciones - (df_suscripciones['has_13'].sum() + df_suscripciones['has_14'].sum() + df_suscripciones['has_15'].sum())
        ]
    })
    
    if total_suscripciones > 0:
        tabla_etnias['Porcentaje'] = (tabla_etnias['Total Suscripciones'] / total_suscripciones * 100).round(2)
        tabla_etnias['Porcentaje'] = tabla_etnias['Porcentaje'].apply(agregar_porcentaje)
    else:
        tabla_etnias['Porcentaje'] = "0%"
    
    return tabla_etnias

def crear_tabla_shades(df_suscripciones):
    """
    Crea una tabla con el total de suscripciones por shade
    """
    total_suscripciones = len(df_suscripciones)
    
    datos_shades = []
    for item_id, shade_name in shades.items():
        col_name = f'shade_{item_id[-4:]}'
        count = df_suscripciones[col_name].sum()
        datos_shades.append({
            'Shade': shade_name,
            'Total Suscripciones': count
        })
    
    tabla_shades = pd.DataFrame(datos_shades)
    
    if total_suscripciones > 0:
        tabla_shades['Porcentaje'] = (tabla_shades['Total Suscripciones'] / total_suscripciones * 100).round(2)
        tabla_shades['Porcentaje'] = tabla_shades['Porcentaje'].apply(agregar_porcentaje)
    else:
        tabla_shades['Porcentaje'] = "0%"
    
    # Ordenar por total de suscripciones (descendente)
    tabla_shades = tabla_shades.sort_values('Total Suscripciones', ascending=False)
    
    return tabla_shades

def analizar_cancelaciones_por_razon(df):
    """
    Analiza las cancelaciones por razón procesada (con etnias)
    """
    # Procesar las razones
    df['razon_procesada'] = df['reason'].apply(procesar_razon)
    
    # Agrupar por razón procesada
    resultado = df.groupby('razon_procesada').agg(
        total_cancelaciones=('id', 'count'),
        suscripciones_unicas=('subscriptionId', 'nunique'),
        caucasian=('has_13', 'sum'),
        african=('has_14', 'sum'),
        asian=('has_15', 'sum')
    ).reset_index()
    
    # Calcular porcentajes
    total_cancelaciones = resultado['total_cancelaciones'].sum()
    
    # Porcentaje del total
    if total_cancelaciones > 0:
        resultado['porcentaje_cancelaciones'] = (resultado['total_cancelaciones'] / total_cancelaciones * 100).round(2)
    else:
        resultado['porcentaje_cancelaciones'] = 0
    
    # Porcentaje para cada segmentación (basado en el total de cada categoría)
    total_caucasian = resultado['caucasian'].sum()
    total_african = resultado['african'].sum()
    total_asian = resultado['asian'].sum()
    
    if total_caucasian > 0:
        resultado['porcentaje_caucasian'] = (resultado['caucasian'] / total_caucasian * 100).round(2)
    else:
        resultado['porcentaje_caucasian'] = 0
        
    if total_african > 0:
        resultado['porcentaje_african'] = (resultado['african'] / total_african * 100).round(2)
    else:
        resultado['porcentaje_african'] = 0
        
    if total_asian > 0:
        resultado['porcentaje_asian'] = (resultado['asian'] / total_asian * 100).round(2)
    else:
        resultado['porcentaje_asian'] = 0
    
    # Agregar símbolo % a las columnas de porcentaje
    columnas_porcentaje = ['porcentaje_cancelaciones', 'porcentaje_caucasian', 'porcentaje_african', 'porcentaje_asian']
    for col in columnas_porcentaje:
        resultado[col] = resultado[col].apply(agregar_porcentaje)
    
    # Ordenar por total de cancelaciones (descendente)
    resultado = resultado.sort_values('total_cancelaciones', ascending=False)
    
    return resultado

def analizar_cancelaciones_por_razon_y_shade(df, filtro_etnia=None):
    """
    Analiza las cancelaciones por razón procesada con columnas para cada shade
    Si se especifica filtro_etnia, solo incluye cancelaciones de esa etnia
    """
    # Filtrar por etnia si se especifica
    if filtro_etnia == 13:
        df = df[df['has_13'] == True]
        nombre_etnia = "CAUCASIAN"
    elif filtro_etnia == 14:
        df = df[df['has_14'] == True]
        nombre_etnia = "AFRICAN"
    elif filtro_etnia == 15:
        df = df[df['has_15'] == True]
        nombre_etnia = "ASIAN"
    else:
        nombre_etnia = "TODAS"
    
    # Procesar las razones
    df['razon_procesada'] = df['reason'].apply(procesar_razon)
    
    # Crear columnas booleanas para cada shade
    for item_id, shade_name in shades.items():
        # Crear nombre de columna válido (sin espacios ni caracteres especiales)
        col_name = f"shade_{item_id[-4:]}"
        df[col_name] = df['itemIds'].str.contains(item_id, na=False)
    
    # Agrupar por razón procesada
    agg_dict = {
        'total_cancelaciones': ('id', 'count'),
        'suscripciones_unicas': ('subscriptionId', 'nunique')
    }
    
    # Agregar cada shade al diccionario de agregación
    for item_id in shades.keys():
        col_name = f"shade_{item_id[-4:]}"
        agg_dict[col_name] = (col_name, 'sum')
    
    resultado = df.groupby('razon_procesada').agg(**agg_dict).reset_index()
    
    # Calcular porcentajes
    total_cancelaciones = resultado['total_cancelaciones'].sum()
    
    # Porcentaje del total
    if total_cancelaciones > 0:
        resultado['porcentaje_cancelaciones'] = (resultado['total_cancelaciones'] / total_cancelaciones * 100).round(2)
    else:
        resultado['porcentaje_cancelaciones'] = 0
    
    # Calcular porcentajes para cada shade
    for item_id in shades.keys():
        col_name = f"shade_{item_id[-4:]}"
        total_shade = resultado[col_name].sum()
        
        if total_shade > 0:
            resultado[f'porcentaje_{col_name}'] = (resultado[col_name] / total_shade * 100).round(2)
        else:
            resultado[f'porcentaje_{col_name}'] = 0
    
    # Renombrar las columnas de shades con nombres legibles
    rename_dict = {}
    for item_id, shade_name in shades.items():
        col_name = f"shade_{item_id[-4:]}"
        rename_dict[col_name] = shade_name
        rename_dict[f'porcentaje_{col_name}'] = f'porcentaje_{shade_name}'
    
    resultado = resultado.rename(columns=rename_dict)
    
    # Agregar símbolo % a todas las columnas de porcentaje
    columnas_porcentaje = ['porcentaje_cancelaciones'] + [f'porcentaje_{shade_name}' for shade_name in shades.values()]
    for col in columnas_porcentaje:
        if col in resultado.columns:
            resultado[col] = resultado[col].apply(agregar_porcentaje)
    
    # Ordenar por total de cancelaciones (descendente)
    resultado = resultado.sort_values('total_cancelaciones', ascending=False)
    
    return resultado, nombre_etnia

def ajustar_ancho_columnas(archivo_excel):
    """
    Ajusta automáticamente el ancho de las columnas en un archivo Excel
    """
    try:
        # Cargar el workbook
        wb = load_workbook(archivo_excel)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Ajustar el ancho de cada columna basado en el contenido
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            # Calcular la longitud del contenido
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Ajustar el ancho de la columna (agregar un poco de padding)
                adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres de ancho
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar los cambios
        wb.save(archivo_excel)
        print(f"✓ Ancho de columnas ajustado para {archivo_excel}")
        
    except Exception as e:
        print(f"✗ Error ajustando el ancho de columnas para {archivo_excel}: {e}")

# Ejemplo de uso
def main(startDate, endDate):
    # 1. Obtener datos de la consulta SQL
    # Crear la lista de itemIds para la consulta (solo 30ml)
    item_ids = list(shades.keys())
    item_ids_str = "', '".join(item_ids)
    
    query = f"""
    SELECT  
        fc.id,
        fc.subscriptionId,
        fc.reason,
        fc.createdAt,
        GROUP_CONCAT(DISTINCT fso.itemId) AS itemIds,
        GROUP_CONCAT(DISTINCT fo.order_number) AS orderNumbers,
        (
            SELECT psoi.additionalFields->>"$.diagnostic"
            FROM prod_sales_and_subscriptions.sales_order_items psoi 
            WHERE psoi.subscriptionId = fc.subscriptionId 
            LIMIT 1
        ) AS additionalFields
    FROM bi.fact_cancellations fc
    JOIN bi.fact_orders fo ON fc.subscriptionId = fo.subscription_id
    JOIN bi.fact_sales_order_items fso ON fo.id = fso.salesOrderId
    WHERE fso.itemId IN ('{item_ids_str}')
      AND fc.createdAt BETWEEN '{startDate}' AND '{endDate}'
    GROUP BY fc.id, fc.subscriptionId, fc.reason, fc.createdAt;
    """
    
    # Ejecutar la consulta (asumiendo que tienes una función execute_query)
    df = execute_query(query)
    
    # Procesar los campos de diagnóstico
    print("Procesando datos de diagnóstico...")
    df['diagnostic_values'] = df['additionalFields'].apply(extract_diagnostic_values)
    
    # Crear columnas booleanas para cada valor
    df['has_13'] = df['diagnostic_values'].apply(lambda x: 13 in x)
    df['has_14'] = df['diagnostic_values'].apply(lambda x: 14 in x)
    df['has_15'] = df['diagnostic_values'].apply(lambda x: 15 in x)
    
    # 2. Obtener datos de suscripciones activas
    print("Obteniendo datos de suscripciones activas...")
    df_suscripciones = obtener_suscripciones_activas(startDate, endDate)
    
    # 3. Crear tablas adicionales
    print("Creando tablas de suscripciones...")
    tabla_etnias = crear_tabla_etnias(df_suscripciones)
    tabla_shades = crear_tabla_shades(df_suscripciones)
    
    # 4. Procesar los datos para las páginas
    print("Procesando datos de cancelaciones por razón (con etnias)...")
    df_por_razon = analizar_cancelaciones_por_razon(df)
    
    print("Procesando datos de cancelaciones por razón (con shades)...")
    df_por_razon_y_shade, _ = analizar_cancelaciones_por_razon_y_shade(df)
    
    print("Procesando datos de cancelaciones por razón (CAUCASIAN)...")
    df_caucasian, _ = analizar_cancelaciones_por_razon_y_shade(df, 13)
    
    print("Procesando datos de cancelaciones por razón (AFRICAN)...")
    df_african, _ = analizar_cancelaciones_por_razon_y_shade(df, 14)
    
    print("Procesando datos de cancelaciones por razón (ASIAN)...")
    df_asian, _ = analizar_cancelaciones_por_razon_y_shade(df, 15)
    
    # 5. Guardar en Excel
    nombre_archivo = f"analisis_cancelaciones_{startDate}_to_{endDate}.xlsx"
    
    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        # Hoja 1: Por Razon (Etnias)
        df_por_razon.to_excel(writer, sheet_name='Por Razon (Etnias)', index=False, startrow=0)
        tabla_etnias.to_excel(writer, sheet_name='Por Razon (Etnias)', index=False, startrow=len(df_por_razon) + 3)
        
        # Hoja 2: Por Razon (Shades) - Múltiples tablas
        start_row = 0
        
        # Tabla principal (Todas las etnias)
        df_por_razon_y_shade.to_excel(writer, sheet_name='Por Razon (Shades)', index=False, startrow=start_row)
        start_row += len(df_por_razon_y_shade) + 3
        
        # Tabla CAUCASIAN
        writer.sheets['Por Razon (Shades)'].cell(row=start_row, column=1, value="CAUCASIAN (13)")
        start_row += 1
        df_caucasian.to_excel(writer, sheet_name='Por Razon (Shades)', index=False, startrow=start_row)
        start_row += len(df_caucasian) + 3
        
        # Tabla AFRICAN
        writer.sheets['Por Razon (Shades)'].cell(row=start_row, column=1, value="AFRICAN (14)")
        start_row += 1
        df_african.to_excel(writer, sheet_name='Por Razon (Shades)', index=False, startrow=start_row)
        start_row += len(df_african) + 3
        
        # Tabla ASIAN
        writer.sheets['Por Razon (Shades)'].cell(row=start_row, column=1, value="ASIAN (15)")
        start_row += 1
        df_asian.to_excel(writer, sheet_name='Por Razon (Shades)', index=False, startrow=start_row)
        start_row += len(df_asian) + 3
        
        # Tabla de suscripciones activas por shade
        writer.sheets['Por Razon (Shades)'].cell(row=start_row, column=1, value="Suscripciones Activas por Shade")
        start_row += 1
        tabla_shades.to_excel(writer, sheet_name='Por Razon (Shades)', index=False, startrow=start_row)
    
    # 6. Ajustar ancho de columnas
    ajustar_ancho_columnas(nombre_archivo)
    
    print(f"\n✅ Análisis completado. Archivo guardado como: {nombre_archivo}")
    print(f"Total de cancelaciones: {df_por_razon['total_cancelaciones'].sum()}")
    print(f"Total de suscripciones únicas canceladas: {df_por_razon['suscripciones_unicas'].sum()}")
    print(f"Total de suscripciones activas: {len(df_suscripciones)}")
    
    # Estadísticas de diagnóstico
    print(f"Cancelaciones CAUCASIAN (13): {df['has_13'].sum()}")
    print(f"Cancelaciones AFRICAN (14): {df['has_14'].sum()}")
    print(f"Cancelaciones ASIAN (15): {df['has_15'].sum()}")
    print(f"Cancelaciones sin diagnóstico: {len(df) - (df['has_13'].sum() + df['has_14'].sum() + df['has_15'].sum())}")
    
    print("\n📋 Top 5 razones de cancelación (Etnias):")
    print(df_por_razon[['razon_procesada', 'total_cancelaciones', 'porcentaje_cancelaciones']].head())
    
    print("\n📋 Distribución de suscripciones activas por etnia:")
    print(tabla_etnias)

# Ejecutar para los períodos deseados
if __name__ == "__main__":
    main('2025-01-01', '2025-04-01')
    # main('2025-05-01', '2025-08-30')