import pandas as pd
import tkinter as tk
from tkinter import Label, Button, messagebox, Listbox, END, MULTIPLE, Frame
from tkcalendar import Calendar
from modules.database_queries import execute_query

# Diccionario de productos actualizado
product_dict = {
     'IT00000000000000000000000000000001': 'Big Shipper: Small Gloves',
        'IT00000000000000000000000000000002': '30ml Colorant - Medium Brown',
        'IT00000000000000000000000000000005': 'Small Shipper: X-Large Gloves',
        'IT00000000000000000000000000000006': '30ml Colorant - Jet-Black',
        'IT00000000000000000000000000000007': '30ml Colorant - Black',
        'IT00000000000000000000000000000008': '30ml Colorant - Soft-Black',
        'IT00000000000000000000000000000009': '30ml Colorant - Dark Brown',
        'IT00000000000000000000000000000010': '30ml Colorant - Medium-Dark Brown',
        'IT00000000000000000000000000000011': '30ml Colorant - Medium-Light Brown',
        'IT00000000000000000000000000000012': '30ml Colorant - Light Brown',
        'IT00000000000000000000000000000013': '30ml Colorant - Dark Blond',
        'IT00000000000000000000000000000014': '30ml Colorant - Warm- Dark Blond',
        'IT00000000000000000000000000000015': '30ml Colorant - Cool- Medium Blond',
        'IT00000000000000000000000000000016': '30ml Colorant - Warm- Medium Blond',
        'IT00000000000000000000000000000017': '30ml Colorant - Cool- Light Blond',
        'IT00000000000000000000000000000018': '30ml Colorant - Warm- Light Blond',
        'IT00000000000000000000000000000019': '30ml Colorant - Lightest Blond',
        'IT00000000000000000000000000000020': '30ml Colorant - Dark Auburn',
        'IT00000000000000000000000000000021': '30ml Colorant - Light Auburn',
        'IT00000000000000000000000000000022': '30ml Developer - 20 Vol',
        'IT00000000000000000000000000000023': '30ml Developer - 10 Vol',
        'IT00000000000000000000000000000024': 'Tray with Spatula',
        'IT00000000000000000000000000000025': '7 Min: Short Beard Brush',
        'IT00000000000000000000000000000026': '10 Min: Short Beard Brush',
        'IT00000000000000000000000000000028': '10 Min: Medium Beard Brush',
        'IT00000000000000000000000000000030': '10 Min: Long Beard Brush',
        'IT00000000000000000000000000000031': 'Small Shipper: Normal Hair Treatment',
        'IT00000000000000000000000000000032': 'Small Shipper: Thin Hair Treatment',
        'IT00000000000000000000000000000033': 'Small Shipper: Ultra Dry Hair Treatment',
        'IT00000000000000000000000000000034': 'Small Shipper: Damaged Hair Treatment',
        'IT00000000000000000000000000000035': 'Small Shipper: Coily Hair Treatment',
        'IT00000000000000000000000000000036': 'Small Shipper: Moisturizing',
        'IT00000000000000000000000000000037': 'Small Shipper: Sensitive',
        'IT00000000000000000000000000000038': 'Small Shipper: Energizing',
        'IT00000000000000000000000000000039': 'Small Shipper',
        'IT00000000000000000000000000000040': 'Big Shipper',
        'IT00000000000000000000000000000041': 'Small Shipper: Small Gloves',
        'IT00000000000000000000000000000042': 'Big Shipper: Medium Gloves',
        'IT00000000000000000000000000000043': 'Big Shipper: Large Gloves',
        'IT00000000000000000000000000000044': 'Big Shipper: X-Large Gloves',
        'IT00000000000000000000000000000045': '45ml Colorant - Jet-Black',
        'IT00000000000000000000000000000046': '45ml Colorant - Black',
        'IT00000000000000000000000000000047': '45ml Colorant - Soft-Black',
        'IT00000000000000000000000000000048': '45ml Colorant - Dark Brown',
        'IT00000000000000000000000000000049': '45ml Colorant - Medium-Dark Brown',
        'IT00000000000000000000000000000050': '45ml Colorant - Medium Brown',
        'IT00000000000000000000000000000051': '45ml Colorant - Medium-Light Brown',
        'IT00000000000000000000000000000052': '45ml Colorant: Light Brown',
        'IT00000000000000000000000000000053': '45ml Colorant - Dark Blond',
        'IT00000000000000000000000000000054': '45ml Colorant - Warm-Dark Blond',
        'IT00000000000000000000000000000055': '45ml Colorant - Cool-Medium Blond',
        'IT00000000000000000000000000000056': '45ml Colorant - Warm-Medium Blond',
        'IT00000000000000000000000000000057': '45ml Colorant - Cool - Light Blond',
        'IT00000000000000000000000000000058': '45ml Colorant - Warm - Light Blond',
        'IT00000000000000000000000000000059': '45ml Colorant - Lightest Blond',
        'IT00000000000000000000000000000060': '45ml Colorant - Dark Auburn',
        'IT00000000000000000000000000000061': '45ml Colorant - Light Auburn',
        'IT00000000000000000000000000000062': '45ml Developer - 30 Vol',
        'IT00000000000000000000000000000063': '45ml Developer - 20 Vol',
        'IT00000000000000000000000000000064': '5 Min: Big Shipper - Brush and Tray',
        'IT00000000000000000000000000000065': '7 Min: Big Shipper - Brush and Tray',
        'IT00000000000000000000000000000066': '10 Min: Big Shipper - Brush and Tray',
        'IT00000000000000000000000000000067': '5 Min: Big Shipper - Brush+Tray+Spatula',
        'IT00000000000000000000000000000068': '7 Min: Big Shipper - Brush+Tray+Spatula',
        'IT00000000000000000000000000000069': '10 Min: Big Shipper - Brush+Tray+Spatula',
        'IT00000000000000000000000000000070': 'Big Shipper: Normal Hair Treatment',
        'IT00000000000000000000000000000071': 'Big Shipper: Thin Hair Treatment',
        'IT00000000000000000000000000000072': 'Big Shipper: Ultra Dry Hair Treatment',
        'IT00000000000000000000000000000073': 'Big Shipper: Damaged Hair Treatment',
        'IT00000000000000000000000000000074': 'Big Shipper: 4C / Afro-Textured Hair',
        'IT00000000000000000000000000000075': 'Box 2 MIX Empty - Big Shipper',
        'IT00000000000000000000000000000076': 'Box 2 MIX Pre-Packed - Small Shipper',
        'IT00000000000000000000000000000004': 'Small Shipper: Large Gloves',
        'IT00000000000000000000000000000003': 'Small Shipper: Medium Gloves',
        'IT00000000000000000000000000000029': '7 Min: Long Beard Brush',
        'IT00000000000000000000000000000027': '7 Min: Medium Beard Brush',
        'IT00000000000000000000000000000103': 'Beard Scrub 7ml Sensitive',
        'IT00000000000000000000000000000104': 'Beard Scrub 7ml Energizing',
    'IT00000000000000000000000000000082': 'Lightest Blond Hair and Beard Dye',
    'IT00000000000000000000000000000083': 'Medium Blond Hair and Beard Dye',
    'IT00000000000000000000000000000084': 'Auburn Hair and Beard Dye',
    'IT00000000000000000000000000000085': 'Light Brown Hair and Beard Dye',
    'IT00000000000000000000000000000086': 'Medium Brown Hair and Beard Dye',
    'IT00000000000000000000000000000087': 'Dark Brown Hair and Beard Dye',
    'IT00000000000000000000000000000088': 'Natural Black Hair and Beard Dye',
    'IT00000000000000000000000000000089': 'Jet Black Hair and Beard Dye',
    'IT00000000000000000000000000000098': 'Dark Blond Hair and Beard Dye',
    'IT00000000000000000000000000000102': 'Black Hair and Beard Dye',
    'IT00000000000000000000000000000105': 'Jet Black Hair and Beard Dye FBM',
    'IT00000000000000000000000000000106': 'Dark Brown Hair and Beard Dye FBM',
}

def open_rebuy_date_selector():
    """Interfaz para seleccionar fechas de análisis"""
    start_date = None
    end_date = None
    
    def get_dates():
        nonlocal start_date, end_date
        start_date = cal_start.get_date()
        end_date = cal_end.get_date()
        
        if not start_date or not end_date:
            messagebox.showerror("Error", "Selecciona ambas fechas.")
            return
            
        # Convertir a formato YYYY-MM-DD
        start_date = pd.to_datetime(start_date).strftime('%Y-%m-%d')
        end_date = pd.to_datetime(end_date).strftime('%Y-%m-%d')
        
        root.quit()
        root.destroy()

    # Configuración de la ventana
    root = tk.Tk()
    root.title("Selección de Rango de Fechas")
    
    Label(root, text="Fecha de inicio:").pack(pady=5)
    cal_start = Calendar(root, selectmode='day', date_pattern='yyyy-mm-dd')
    cal_start.pack(pady=5)

    Label(root, text="Fecha de fin:").pack(pady=5)
    cal_end = Calendar(root, selectmode='day', date_pattern='yyyy-mm-dd')
    cal_end.pack(pady=5)

    Button(root, text="Generar Reporte", command=get_dates).pack(pady=20)
    
    root.mainloop()
    return start_date, end_date

def select_report_type():
    """Interfaz para seleccionar tipo de reporte"""
    report_types = []
    
    def get_selection():
        nonlocal report_types
        report_types = []
        if var_total.get():
            report_types.append('total')
        if var_combinations.get():
            report_types.append('combinations')
        
        if not report_types:
            messagebox.showerror("Error", "Selecciona al menos un tipo de reporte")
            return
            
        root.quit()
        root.destroy()

    root = tk.Tk()
    root.title("Selección de Tipo de Reporte")
    
    var_total = tk.BooleanVar()
    var_combinations = tk.BooleanVar()
    
    tk.Label(root, text="Selecciona el tipo de reporte:").pack(pady=10)
    tk.Checkbutton(root, text="Reporte Total (por producto individual)", variable=var_total).pack(anchor='w')
    tk.Checkbutton(root, text="Reporte de Combinaciones", variable=var_combinations).pack(anchor='w')
    
    Button(root, text="Continuar", command=get_selection).pack(pady=20)
    
    root.mainloop()
    return report_types

def select_product_combinations():
    """Interfaz para seleccionar combinaciones de productos"""
    display_names = sorted(set(product_dict.values()))
    selected_combinations = []
    
    def add_combination():
        selected = [display_names[i] for i in listbox.curselection()]
        if len(selected) >= 2:
            combo_name = " + ".join(selected)
            combinations_listbox.insert(END, combo_name)
            selected_combinations.append(selected)
        else:
            messagebox.showwarning("Advertencia", "Selecciona al menos 2 productos para una combinación")
    
    def remove_combination():
        selected = combinations_listbox.curselection()
        if selected:
            combinations_listbox.delete(selected[0])
            selected_combinations.pop(selected[0])
    
    def finish_selection():
        root.quit()
        root.destroy()
    
    root = tk.Tk()
    root.title("Seleccionar Combinaciones de Productos")
    
    # Frame para selección de productos
    selection_frame = Frame(root)
    selection_frame.pack(pady=10, padx=10, fill=tk.X)
    
    Label(selection_frame, text="Productos disponibles:").pack()
    listbox = Listbox(selection_frame, selectmode=MULTIPLE, height=10, width=50)
    listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
    scrollbar = tk.Scrollbar(selection_frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    listbox.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=listbox.yview)
    
    for product in display_names:
        listbox.insert(END, product)
    
    # Frame para combinaciones seleccionadas
    combo_frame = Frame(root)
    combo_frame.pack(pady=10, padx=10, fill=tk.X)
    
    Label(combo_frame, text="Combinaciones seleccionadas:").pack()
    combinations_listbox = Listbox(combo_frame, height=5, width=50)
    combinations_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
    # Botones de acción
    button_frame = Frame(root)
    button_frame.pack(pady=10)
    
    Button(button_frame, text="Añadir Combinación", command=add_combination).pack(side=tk.LEFT, padx=5)
    Button(button_frame, text="Eliminar Combinación", command=remove_combination).pack(side=tk.LEFT, padx=5)
    Button(button_frame, text="Finalizar Selección", command=finish_selection).pack(side=tk.LEFT, padx=5)
    
    root.mainloop()
    return selected_combinations

def process_individual_products(df):
    """Procesa el reporte para productos individuales con la nueva lógica"""
    # Filtrar solo los productos que están en nuestro diccionario
    df = df[df['itemId'].isin(product_dict.keys())]
    
    # Mapear los itemId a nombres de productos
    df['product_name'] = df['itemId'].map(product_dict)
    
    # Obtener todas las órdenes de los usuarios
    user_orders = df.groupby('customer_id')['id'].nunique().reset_index()
    user_orders.columns = ['customer_id', 'total_orders']
    
    # Separar usuarios con solo 1 orden vs múltiples órdenes
    single_order_users = user_orders[user_orders['total_orders'] == 1]['customer_id']
    multi_order_users = user_orders[user_orders['total_orders'] > 1]['customer_id']
    
    # Obtener productos de primera orden para cada usuario
    first_orders = df.sort_values(['customer_id', 'created_at']).groupby('customer_id').first().reset_index()
    first_order_products = df.merge(first_orders[['customer_id', 'id']], 
                                  on=['customer_id', 'id'])
    
    # Calcular usuarios que recompraron (tienen >1 orden)
    reordered = first_order_products[first_order_products['customer_id'].isin(multi_order_users)]
    reordered_count = reordered['product_name'].value_counts().reset_index()
    reordered_count.columns = ['product_name', 'usuarios_recompraron']
    
    # Calcular usuarios que NO recompraron (solo 1 orden)
    not_reordered = first_order_products[first_order_products['customer_id'].isin(single_order_users)]
    not_reordered_count = not_reordered['product_name'].value_counts().reset_index()
    not_reordered_count.columns = ['product_name', 'usuarios_no_recompraron']
    
    # Combinar resultados
    result = pd.merge(reordered_count, not_reordered_count, on='product_name', how='outer').fillna(0)
    result['total_usuarios'] = result['usuarios_recompraron'] + result['usuarios_no_recompraron']
    result['porcentaje_recompra'] = (result['usuarios_recompraron'] / result['total_usuarios'] * 100).round(2)
    
    # Ordenar y formatear
    result = result.sort_values('product_name', ascending=True)
    result = result[['product_name', 'usuarios_recompraron', 'usuarios_no_recompraron', 
                    'total_usuarios', 'porcentaje_recompra']]
    
    return result

def process_product_combinations(df, combinations):
    """Procesa combinaciones de productos con la nueva lógica"""
    # Filtrar solo los productos que están en nuestro diccionario
    df = df[df['itemId'].isin(product_dict.keys())]
    df['product_name'] = df['itemId'].map(product_dict)
    
    # Obtener todas las órdenes de los usuarios
    user_orders = df.groupby('customer_id')['id'].nunique().reset_index()
    user_orders.columns = ['customer_id', 'total_orders']
    
    # Separar usuarios con solo 1 orden vs múltiples órdenes
    single_order_users = user_orders[user_orders['total_orders'] == 1]['customer_id']
    multi_order_users = user_orders[user_orders['total_orders'] > 1]['customer_id']
    
    # Obtener productos de primera orden para cada usuario
    first_orders = df.sort_values(['customer_id', 'created_at']).groupby('customer_id').first().reset_index()
    first_order_products = df.merge(first_orders[['customer_id', 'id']], 
                                  on=['customer_id', 'id'])
    
    # Preparar resultados
    results = []
    
    for combo in combinations:
        combo_name = " + ".join(sorted(combo))
        
        # Usuarios que tuvieron TODOS los productos de la combinación en su primera orden
        users_with_combo = first_order_products.groupby('customer_id').filter(
            lambda x: all(p in x['product_name'].values for p in combo)
        )['customer_id'].unique()
        
        if len(users_with_combo) > 0:
            # Usuarios que recompraron (tienen >1 orden)
            reordered = len(set(users_with_combo) & set(multi_order_users))
            
            # Usuarios que NO recompraron (solo 1 orden)
            not_reordered = len(set(users_with_combo) & set(single_order_users))
            
            total = reordered + not_reordered
            rebuy_percentage = round(reordered / total * 100, 2) if total > 0 else 0  # Línea corregida
            
            results.append({
                'product_name': combo_name,
                'usuarios_recompraron': reordered,
                'usuarios_no_recompraron': not_reordered,
                'total_usuarios': total,
                'porcentaje_recompra': rebuy_percentage
            })
    
    # Crear DataFrame con los resultados
    result = pd.DataFrame(results)
    
    # Ordenar por nombre de combinación
    if not result.empty:
        result = result.sort_values('product_name', ascending=True)
    
    return result

def main():
    # 1. Obtener fechas del selector
    start_date, end_date = open_rebuy_date_selector()
    
    if not start_date or not end_date:
        return
    
    # 2. Seleccionar tipo de reporte
    report_types = select_report_type()
    
    if not report_types:
        return
    
    # 3. Si se seleccionó combinaciones, obtener las combinaciones
    combinations = []
    if 'combinations' in report_types:
        combinations = select_product_combinations()
        if not combinations:
            messagebox.showwarning("Advertencia", "No se seleccionaron combinaciones. No se generará reporte de combinaciones.")
            report_types.remove('combinations')
            if not report_types:
                return
    
    # 4. Obtener datos con la nueva consulta que incluye created_at
    query_all_orders = """
    WITH ordenes_con_beard AS (
        SELECT DISTINCT fo.id
        FROM bi.fact_orders fo
        JOIN bi.fact_sales_order_items oi ON fo.id = oi.salesOrderId
        WHERE oi.category = 'IG00000000000000000000000000000029'
    )
    SELECT fo.customer_id, fo.id, oi.itemId, fo.created_at
    FROM bi.fact_orders fo
    JOIN bi.fact_sales_order_items oi ON fo.id = oi.salesOrderId
    WHERE fo.status != 'CANCELLED'
    AND fo.order_plan = 'OTO'
    AND fo.recurrent = 0
    AND fo.id IN (SELECT id FROM ordenes_con_beard)
    """
    main_df = execute_query(query_all_orders)
    
    # 5. Obtener datos de primera orden
    query_first_orders = f"""
    SELECT fo.customer_id
    FROM bi.fact_orders fo
    JOIN bi.fact_sales_order_items oi ON fo.id = oi.salesOrderId
    WHERE fo.status != 'CANCELLED'
    AND fo.order_plan = 'OTO'
    AND fo.recurrent = 0
    AND fo.is_first_order = 1
    AND fo.created_at BETWEEN '{start_date}' AND '{end_date}'
    """
    first_time_buyers_df = execute_query(query_first_orders)
    
    if not main_df.empty and not first_time_buyers_df.empty:
        # Filtrar solo usuarios que están en el listado de primera compra
        valid_users = first_time_buyers_df['customer_id'].unique()
        filtered_df = main_df[main_df['customer_id'].isin(valid_users)]
        
        # Procesar los reportes seleccionados
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        if 'total' in report_types:
            individual_result = process_individual_products(filtered_df)
            ws1 = wb.create_sheet(title="Productos_Individuales")
            
            # Escribir los datos
            for r in dataframe_to_rows(individual_result, index=False, header=True):
                ws1.append(r)
            
            # Aplicar formato
            for row in ws1.iter_rows(min_row=2, min_col=5, max_col=5):
                for cell in row:
                    cell.number_format = '0.00%'
                    if cell.value is not None:
                        cell.value = cell.value / 100
            
            ws1.freeze_panes = 'A2'
            
            # Ajustar anchos
            for column in ws1.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                ws1.column_dimensions[column[0].column_letter].width = max_length + 2
        
        if 'combinations' in report_types and combinations:
            combo_result = process_product_combinations(filtered_df, combinations)
            ws2 = wb.create_sheet(title="Combinaciones")
            
            for r in dataframe_to_rows(combo_result, index=False, header=True):
                ws2.append(r)
                
            for row in ws2.iter_rows(min_row=2, min_col=5, max_col=5):
                for cell in row:
                    cell.number_format = '0.00%'
                    if cell.value is not None:
                        cell.value = cell.value / 100
            
            ws2.freeze_panes = 'A2'
            
            for column in ws2.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                ws2.column_dimensions[column[0].column_letter].width = max_length + 2
        
        # Eliminar hoja por defecto y guardar
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        wb.save(f'reporte_recompras_first_order_{start_date}_a_{end_date}.xlsx')
        print("Reporte generado exitosamente con la nueva metodología")
    else:
        print("No se encontraron datos para generar el reporte")

if __name__ == "__main__":
    main()