import pandas as pd
import json
import tkinter as tk
from tkinter import Label, Button, messagebox, Listbox, END, MULTIPLE, Frame
from tkcalendar import Calendar
from modules.database_queries import execute_query
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from ast import literal_eval
import base64

# Diccionarios para mapeo de variables y valores
variables_dict = {
    1: "HEAD OF HAIR",
    2: "HAIR CONDITION",
    3: "TEXTURE",
    4: "HAIR TYPE",
    5: "HAIR STYLE",
    6: "AGE",
    7: "GLOVE SIZE",
    8: "NATURAL COLOR",
    9: "NATURAL SHADE",
    10: "CURRENT GRAY",
    11: "GRAY CONCENTRATION",
    12: "GRAY DIFFICULTY",
    13: "STUBBORN GRAYS",
    14: "QUIZ TYPE",
    15: "PRIMARY OBJECTIVE",
    16: "STAY THE SAME",
    17: "BEARD LENGTH",
    18: "BEARD THICK",
    19: "BEARD SCRUB",
    20: "DESIRED COLOR",
    21: "DESIRED SHADE",
    24: "DESIRED COVERAGE",
    27: "USER FIRST NAME",
    29: "USER EMAIL",
    30: "SALT & PEPPER LOOK",
    31: "VELOCITY",
    32: "COLOR INCREASE",
    33: "SHIPPER",
    34: "PURCHASE NUMBER",
    35: "USER LAST NAME",
    36: "EXPERIENCE WITH COLOR",
    37: "COLORED HAIR",
    39: "HAIR TREATMENT",
    40: "B - EXPERIENCE WITH COLOR",
    41: "SKIN REACTION",
    42: "CONTACT CUSTOMER SERVICE",
    43: "PURCHASE PLAN",
    44: "SUBSCRIPTION STARTING DATE",
    45: "USE DEVELOPER FROM ITALY"
}

values_dict = {
    4: "BALDING",
    5: "FULL",
    6: "STARTING TO THIN",
    7: "ULTRA DRY HAIR",
    8: "NORMAL",
    9: "DAMAGED HAIR",
    10: "THIN / FINE HAIR",
    11: "NORMAL HAIR",
    12: "THICK / COARSE",
    13: "CAUCASIAN",
    14: "AFRICAN",
    15: "ASIAN",
    16: "HS1",
    17: "HS2",
    18: "HS3",
    19: "HS4",
    20: "HS5",
    21: "HS6",
    22: "HS7",
    23: "HS8",
    24: "< 30",
    25: "30s",
    26: "40s",
    27: "50s",
    28: "60",
    29: "S",
    30: "M",
    31: "L",
    32: "XL",
    33: "BLACK",
    34: "DARK BROWN",
    35: "LIGHT BROWN",
    36: "RED",
    37: "DARK BLOND",
    38: "LIGHT BLOND",
    39: "1.7",
    40: "1",
    41: "2",
    42: "3.1",
    43: "4.01",
    44: "6.31",
    45: "5.13",
    46: "6.12",
    47: "7.01",
    48: "7.31",
    49: "8.01",
    50: "8.31",
    51: "9.12",
    52: "9.31",
    53: "10.03",
    54: "5.04",
    55: "7.04",
    56: "30% OR LESS",
    57: "30%-50%",
    58: "50%-75%",
    59: "NONE",
    60: "MORE THAN 75%",
    61: "SCATTERED",
    62: "CONCENTRATED",
    63: "EASY",
    64: "DIFFICULT",
    65: "DON'T KNOW",
    66: "YES",
    67: "HAIR",
    68: "BEARD",
    69: "NO",
    70: "COVER GRAY & MATCH SHADE",
    71: "CHANGE COLOR",
    72: "YES",
    73: "NO",
    74: "SHORT",
    75: "MEDIUM",
    76: "LONG",
    77: "FULL",
    78: "AVERAGE",
    79: "PATCHY",
    80: "MOISTURIZING",
    81: "ENERGIZING",
    82: "SENSITIVE",
    92: "FULL",
    93: "PARTIAL",
    94: "TARGETED / TOUCH UP",
    95: "FULL COVERAGE - IMMEDIATE",
    96: "FULL COVERAGE - GRADUAL",
    97: "SALT & PEPPER TARGETED",
    98: "SALT & PEPPER BLENDED",
    99: "30% OR LESS",
    100: "30%-50%",
    101: "50%-75%",
    102: "MORE THAN 75%",
    104: "INCREASE",
    105: "DECREASE",
    106: "SAME",
    107: "SMALL",
    108: "BIG",
    109: "1",
    110: "2",
    111: "3",
    112: "CURRENTLY DYED",
    113: "Highlights**",
    114: "Never colored",
    115: "Regular Hair Dye",
    116: "I've Colored",
    117: "COILY HAIR",
    118: "B - Currently Dyed",
    119: "B - I've colored",
    120: "B - Never colored",
    121: "YES",
    122: "NO",
    123: "CONTACT CUSTOMER SERVICE",
    124: "CONTINUE QUIZ",
    125: "OTO",
    126: "SUBSCRIPTION",
    127: "HAIR_INSTRUCTIONS",
    128: "BEARD_INSTRUCTIONS",
    129: "BOTH_INSTRUCTIONS",
    130: "NORMAL",
    131: "DRY",
    132: "YES",
    133: "NO"
}

def open_rebuy_date_selector():
    """Interfaz para seleccionar fechas de análisis"""
    start_date = None
    end_date = None
    
    def get_dates():
        nonlocal start_date, end_date
        start_date = cal_start.get_date()
        end_date = cal_end.get_date()
        
        if not start_date or not end_date:
            messagebox.showerror("Error", "Selecciona ambas fechas.")
            return
            
        # Convertir a formato YYYY-MM-DD
        start_date = pd.to_datetime(start_date).strftime('%Y-%m-%d')
        end_date = pd.to_datetime(end_date).strftime('%Y-%m-%d')
        
        root.quit()
        root.destroy()

    # Configuración de la ventana
    root = tk.Tk()
    root.title("Selección de Rango de Fechas")
    
    Label(root, text="Fecha de inicio:").pack(pady=5)
    cal_start = Calendar(root, selectmode='day', date_pattern='yyyy-mm-dd')
    cal_start.pack(pady=5)

    Label(root, text="Fecha de fin:").pack(pady=5)
    cal_end = Calendar(root, selectmode='day', date_pattern='yyyy-mm-dd')
    cal_end.pack(pady=5)

    Button(root, text="Generar Reporte", command=get_dates).pack(pady=20)
    
    root.mainloop()
    return start_date, end_date

def select_report_type():
    """Interfaz para seleccionar tipo de reporte"""
    report_types = []
    
    def get_selection():
        nonlocal report_types
        report_types = []
        if var_total.get():
            report_types.append('total')
        if var_combinations.get():
            report_types.append('combinations')
        
        if not report_types:
            messagebox.showerror("Error", "Selecciona al menos un tipo de reporte")
            return
            
        root.quit()
        root.destroy()

    root = tk.Tk()
    root.title("Selección de Tipo de Reporte")
    
    var_total = tk.BooleanVar()
    var_combinations = tk.BooleanVar()
    
    tk.Label(root, text="Selecciona el tipo de reporte:").pack(pady=10)
    tk.Checkbutton(root, text="Reporte Total (por producto individual)", variable=var_total).pack(anchor='w')
    tk.Checkbutton(root, text="Reporte de Combinaciones", variable=var_combinations).pack(anchor='w')
    
    Button(root, text="Continuar", command=get_selection).pack(pady=20)
    
    root.mainloop()
    return report_types

def process_individual_products(df):
    """Procesa el reporte para variables:valores individuales"""
    try:
        # Dividir el itemId en variable y valor
        df[['variable_id', 'value_id']] = df['itemId'].str.split(':', expand=True)
        
        # Convertir variable_id a entero (deberían ser todos numéricos)
        df['variable_id'] = df['variable_id'].astype(int)
        
        # Mapear a descripciones - manejar valores no numéricos
        df['variable_name'] = df['variable_id'].map(variables_dict)
        
        # Para value_name, primero intentamos convertir a int para los valores numéricos
        def get_value_description(value_str):
            try:
                # Si es numérico, mapear con values_dict
                value_int = int(value_str)
                return values_dict.get(value_int, value_str)  # Si no está en el dict, devolver el valor original
            except ValueError:
                # Si no es numérico, devolver el valor directamente
                return value_str
        
        df['value_name'] = df['value_id'].apply(get_value_description)
        
        filtered_df10 = df[df['developer'] == 10]
        filtered_df20 = df[df['developer'] == 20]
        filtered_df0 = df[df['developer'] == 0]

        print(f"Todos: {df.shape[0]}, D10: {filtered_df10.shape[0]}, D20: {filtered_df20.shape[0]}, D0: {filtered_df0.shape[0]}")
        # Combinar en una descripción completa
        df['question_response'] = df['variable_name'] + ": " + df['value_name']
        
        # Resto de la lógica igual pero usando question_response
        # Obtener todas las órdenes de los usuarios
        user_orders = df.groupby('customerId')['order_id'].nunique().reset_index()
        user_orders.columns = ['customerId', 'total_orders']
        
        # Separar usuarios con solo 1 orden vs múltiples órdenes
        single_order_users = user_orders[user_orders['total_orders'] == 1]['customerId']
        multi_order_users = user_orders[user_orders['total_orders'] > 1]['customerId']
        
        # Obtener productos de primera orden para cada usuario
        first_orders = df.sort_values(['customerId', 'createdAt']).groupby('customerId').first().reset_index()
        first_order_products = df.merge(first_orders[['customerId', 'order_id']], 
                                    on=['customerId', 'order_id'])
        
        # Calcular usuarios que recompraron (tienen >1 orden)
        reordered = first_order_products[first_order_products['customerId'].isin(multi_order_users)]
        reordered_count = reordered['question_response'].value_counts().reset_index()
        reordered_count.columns = ['question_response', 'usuarios_recompraron']
        
        # Calcular usuarios que NO recompraron (solo 1 orden)
        not_reordered = first_order_products[first_order_products['customerId'].isin(single_order_users)]
        not_reordered_count = not_reordered['question_response'].value_counts().reset_index()
        not_reordered_count.columns = ['question_response', 'usuarios_no_recompraron']
        
        # Combinar resultados
        result = pd.merge(reordered_count, not_reordered_count, on='question_response', how='outer').fillna(0)
        result['total_usuarios'] = result['usuarios_recompraron'] + result['usuarios_no_recompraron']
        result['porcentaje_recompra'] = (result['usuarios_recompraron'] / result['total_usuarios'] * 100).round(2)
        
        # Ordenar y formatear
        result = result.sort_values('question_response', ascending=True)
        result = result[['question_response', 'usuarios_recompraron', 'usuarios_no_recompraron', 
                        'total_usuarios', 'porcentaje_recompra']]
        
        return result
    
    except Exception as e:
        print(f"Error al procesar datos: {str(e)}")
        # Imprimir filas problemáticas para diagnóstico
        print("Valores problemáticos en value_id:")
        print(df[~df['value_id'].str.isnumeric()]['value_id'].unique())
        raise

def process_product_combinations(df, combinations):
    """Procesa combinaciones de variables:valores"""
    try:
        # Dividir el itemId en variable y valor
        df[['variable_id', 'value_id']] = df['itemId'].str.split(':', expand=True)
        df['variable_id'] = df['variable_id'].astype(int)
        
        # Mapear a descripciones
        df['variable_name'] = df['variable_id'].map(variables_dict)
        
        # Función para manejar valores no numéricos
        def get_value_description(value_str):
            try:
                return values_dict.get(int(value_str), value_str)
            except ValueError:
                return value_str
        
        df['value_name'] = df['value_id'].apply(get_value_description)
        df['response_description'] = df['variable_name'] + ": " + df['value_name']
        
        # Obtener todas las órdenes de los usuarios
        user_orders = df.groupby('customerId')['order_id'].nunique().reset_index()
        user_orders.columns = ['customerId', 'total_orders']
        
        # Separar usuarios con solo 1 orden vs múltiples órdenes
        single_order_users = user_orders[user_orders['total_orders'] == 1]['customerId']
        multi_order_users = user_orders[user_orders['total_orders'] > 1]['customerId']
        
        # Obtener respuestas de primera orden para cada usuario
        first_orders = df.sort_values(['customerId', 'createdAt']).groupby('customerId').first().reset_index()
        first_order_responses = df.merge(first_orders[['customerId', 'order_id']], 
                                      on=['customerId', 'order_id'])
        
        # Preparar resultados
        results = []
        
        for combo in combinations:
            combo_descriptions = []
            valid_combo = True
            
            # Verificar que todos los items de la combinación sean válidos
            for item in combo:
                if ':' not in item:
                    print(f"Item mal formado en combinación: {item}")
                    valid_combo = False
                    break
                
                var, val = item.split(':')
                try:
                    var_desc = variables_dict.get(int(var), f"Variable {var}")
                    val_desc = get_value_description(val)
                    combo_descriptions.append(f"{var_desc}: {val_desc}")
                except ValueError:
                    print(f"Valor no numérico en combinación: {item}")
                    valid_combo = False
                    break
            
            if not valid_combo:
                continue
            
            combo_name = " + ".join(combo_descriptions)
            
            # Usuarios que tuvieron TODOS los items de la combinación
            users_with_combo = first_order_responses.groupby('customerId').filter(
                lambda x: all(p in x['itemId'].values for p in combo)
            )['customerId'].unique()
            
            if len(users_with_combo) > 0:
                reordered = len(set(users_with_combo) & set(multi_order_users))
                not_reordered = len(set(users_with_combo) & set(single_order_users))
                total = reordered + not_reordered
                rebuy_percentage = round(reordered / total * 100, 2) if total > 0 else 0
                
                results.append({
                    'combination': combo_name,
                    'usuarios_recompraron': reordered,
                    'usuarios_no_recompraron': not_reordered,
                    'total_usuarios': total,
                    'porcentaje_recompra': rebuy_percentage
                })
        
        return pd.DataFrame(results).sort_values('combination')
    
    except Exception as e:
        print(f"Error al procesar combinaciones: {str(e)}")
        raise
            

def proccesdata(start_date, end_date, report_types, combinations, main_df):
    
    # 5. Obtener datos de primera orden (igual que antes)
    query_first_orders = f"""
    WITH 
    -- Primero identificamos la primera orden válida de cada cliente
    primeras_ordenes AS (
        SELECT 
            customerId,
            MIN(createdAt) as primera_fecha_orden
        FROM prod_sales_and_subscriptions.sales_orders
        WHERE status != 'CANCELLED'
        AND status != 'PAYMENT_ERROR'
        GROUP BY customerId
    ),

    -- Filtramos solo las primeras órdenes en el rango de fechas
    primeras_ordenes_en_rango AS (
        SELECT customerId
        FROM primeras_ordenes
        WHERE primera_fecha_orden BETWEEN '{start_date}' AND '{end_date}'
    ),

    -- Identificamos órdenes con productos beard (solo órdenes válidas)
    ordenes_con_beard AS (
        SELECT DISTINCT so.id
        FROM prod_sales_and_subscriptions.sales_orders so
        JOIN prod_sales_and_subscriptions.sales_order_items soi ON so.id = soi.salesOrderId
        JOIN prod_sales_and_subscriptions.items_item_groups ig ON soi.itemId = ig.itemId
        WHERE ig.itemGroupId = 'IG00000000000000000000000000000029'
        AND so.status != 'CANCELLED'
        AND so.status != 'PAYMENT_ERROR'
        AND soi.subscriptionId IS NULL
    )

    -- Seleccionamos los clientes que cumplen todas las condiciones
    SELECT DISTINCT po.customerId
    FROM primeras_ordenes_en_rango po
    WHERE EXISTS (
        SELECT 1
        FROM prod_sales_and_subscriptions.sales_orders so
        WHERE so.customerId = po.customerId
        AND so.id IN (SELECT id FROM ordenes_con_beard)
    )
    """
    first_time_buyers_df = execute_query(query_first_orders)
    
    if not main_df.empty and not first_time_buyers_df.empty:
        # Filtrar solo usuarios que están en el listado de primera compra
        valid_users = first_time_buyers_df['customerId'].unique()
        filtered_df = main_df[main_df['customerId'].isin(valid_users)]
        filtered_df10 = filtered_df[filtered_df['developer'] == 10]
        filtered_df20 = filtered_df[filtered_df['developer'] == 20]

        # Procesar los reportes seleccionados
        wb = Workbook()
        
        # Función auxiliar para añadir hojas al Excel
        def add_sheet_to_excel(wb, df, sheet_name, report_types, combinations):
            """Procesa un DataFrame y añade las hojas correspondientes al workbook"""
            if 'total' in report_types:
                individual_result = process_individual_products(df)
                ws = wb.create_sheet(title=f"{sheet_name}_Individuales")
                
                # Escribir los datos
                for r in dataframe_to_rows(individual_result, index=False, header=True):
                    ws.append(r)
                
                # Aplicar formato
                for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
                    for cell in row:
                        cell.number_format = '0.00%'
                        if cell.value is not None:
                            cell.value = cell.value / 100
                
                ws.freeze_panes = 'A2'
                
                # Ajustar anchos
                for column in ws.columns:
                    max_length = max(len(str(cell.value)) for cell in column)
                    ws.column_dimensions[column[0].column_letter].width = max_length + 2
            
            if 'combinations' in report_types and combinations:
                combo_result = process_product_combinations(df, combinations)
                ws = wb.create_sheet(title=f"{sheet_name}_Combinaciones")
                
                for r in dataframe_to_rows(combo_result, index=False, header=True):
                    ws.append(r)
                    
                for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
                    for cell in row:
                        cell.number_format = '0.00%'
                        if cell.value is not None:
                            cell.value = cell.value / 100
                
                ws.freeze_panes = 'A2'
                
                for column in ws.columns:
                    max_length = max(len(str(cell.value)) for cell in column)
                    ws.column_dimensions[column[0].column_letter].width = max_length + 2
        
        # Procesar cada DataFrame en hojas separadas
        if not filtered_df.empty:
            add_sheet_to_excel(wb, filtered_df, "Todos", report_types, combinations)
        
        if not filtered_df10.empty:
            add_sheet_to_excel(wb, filtered_df10, "Developer10", report_types, combinations)
        
        if not filtered_df20.empty:
            add_sheet_to_excel(wb, filtered_df20, "Developer20", report_types, combinations)
        
        # Eliminar hoja por defecto y guardar
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Verificar que al menos se creó alguna hoja
        if len(wb.sheetnames) > 0:
            wb.save(f'reporte_recompras_diagnostico_{start_date}_a_{end_date}.xlsx')
            print("Reporte generado exitosamente con la nueva metodología")
        else:
            print("No se generaron hojas porque no había datos válidos en ningún DataFrame")
    else:
        print("No se encontraron datos para generar el reporte")

def decode_base64_json(base64_str):
    """Intenta decodificar un string base64 a JSON"""
    try:
        decoded = base64.b64decode(base64_str).decode('utf-8')
        return json.loads(decoded)
    except:
        return None

def process_additional_fields(df):
    """
    Procesa el campo additionalFields para extraer los items de diagnóstico
    Maneja los tres formatos diferentes encontrados
    """
    def extract_items(additional_fields):
        try:
            # Caso 1: additionalFields es None/NaN
            if pd.isna(additional_fields):
                return None
                
            # Caso 2: additionalFields es string (JSON)
            if isinstance(additional_fields, str):
                # Intentar parsear como JSON
                try:
                    af_dict = json.loads(additional_fields)
                except json.JSONDecodeError:
                    # Intentar como literal Python (por si acaso)
                    try:
                        af_dict = literal_eval(additional_fields)
                    except:
                        # Intentar decodificar como base64
                        decoded = decode_base64_json(additional_fields)
                        if decoded is not None:
                            af_dict = decoded
                        else:
                            return None
            # Caso 3: additionalFields ya es un diccionario
            else:
                af_dict = additional_fields
            
            # Verificar si existe diagnostic.values
            if not isinstance(af_dict, dict) or 'diagnostic' not in af_dict:
                return None
            
            diagnostic = af_dict['diagnostic']
            items = []
            
            # Formato 1: values como objeto/diccionario
            if 'values' in diagnostic and isinstance(diagnostic['values'], dict):
                for key, value_data in diagnostic['values'].items():
                    try:
                        variable = str(value_data.get('variable', ''))
                        value = str(value_data.get('value', ''))
                        if variable and value:
                            items.append(f"{variable}:{value}")
                    except:
                        continue
            
            # Formato 2 y 3: values como lista
            elif 'values' in diagnostic and isinstance(diagnostic['values'], list):
                for value_data in diagnostic['values']:
                    try:
                        variable = str(value_data.get('variable', ''))
                        value = str(value_data.get('value', ''))
                        if variable and value:
                            items.append(f"{variable}:{value}")
                    except:
                        continue
            
            return items if items else None
        
        except Exception as e:
            print(f"Error procesando additionalFields: {str(e)}")
            return None
    
    # Aplicar la función a cada fila
    df['items'] = df['additionalFields'].apply(extract_items)
    
    # Eliminar la columna additionalFields
    df = df.drop(columns=['additionalFields'])
    
    return df

def expand_items_to_rows(df):
    """
    Expande cada lista de items en filas individuales
    """
    # Explotar la lista de items para crear filas individuales
    df_expanded = df.explode('items')
    
    # Eliminar filas donde items es None o NaN
    df_expanded = df_expanded[df_expanded['items'].notna()]
    
    # Renombrar la columna items a itemId para consistencia
    df_expanded = df_expanded.rename(columns={'items': 'itemId'})
    
    return df_expanded

def principalQuery():
    # 1. Consulta SQL para obtener los datos
    query = """
    WITH ordenes_con_beard AS (
        SELECT DISTINCT so.id
        FROM prod_sales_and_subscriptions.sales_orders so
        JOIN prod_sales_and_subscriptions.sales_order_items soi ON so.id = soi.salesOrderId
        JOIN prod_sales_and_subscriptions.items_item_groups ig ON soi.itemId = ig.itemId
        WHERE ig.itemGroupId = 'IG00000000000000000000000000000029'
    ),
     ordenes_con_developer20 AS (
        SELECT DISTINCT soi.salesOrderId as id
		FROM prod_sales_and_subscriptions.sales_order_items soi 
        WHERE soi.itemId = 'IT00000000000000000000000000000022'
    ),
    ordenes_con_developer10 AS (
        SELECT DISTINCT soi.salesOrderId as id
		FROM prod_sales_and_subscriptions.sales_order_items soi 
        WHERE soi.itemId = 'IT00000000000000000000000000000023'
    )
    SELECT 
        so.customerId, 
        so.id AS order_id, 
        so.createdAt,
        soi.additionalFields,
	CASE 
		WHEN developer20.id IS NOT NULL THEN 20 
        WHEN developer10.id IS NOT NULL THEN 10 
        ELSE 0
        END as developer
    FROM prod_sales_and_subscriptions.sales_orders so
    JOIN prod_sales_and_subscriptions.sales_order_items soi ON so.id = soi.salesOrderId
    LEFT JOIN ordenes_con_developer20 developer20 ON so.id = developer20.id
    LEFT JOIN ordenes_con_developer10 developer10 ON so.id = developer10.id
    WHERE so.status != 'CANCELLED'
    AND so.status != 'PAYMENT_ERROR'
    AND soi.subscriptionId is null
    AND so.id IN (SELECT id FROM ordenes_con_beard)
    AND soi.itemId LIKE "%100417%"
    """
    
    # 2. Ejecutar consulta
    print("Ejecutando consulta SQL...")
    df = execute_query(query)
    
    if df.empty:
        print("No se encontraron datos para procesar")
        return
    
    # 3. Procesar additionalFields y extraer items
    print("Procesando additionalFields...")
    processed_df = process_additional_fields(df)
    
    # 4. Expandir items en filas individuales
    print("Expandiendo items en filas individuales...")
    expanded_df = expand_items_to_rows(processed_df)
    
    return expanded_df

def main():
    # # 1. Obtener fechas del selector
    # start_date, end_date = open_rebuy_date_selector()
    
    # if not start_date or not end_date:
    #     return
    
    # 2. Seleccionar tipo de reporte
    report_types = select_report_type()
    
    if not report_types:
        return
    
    # 3. Si se seleccionó combinaciones, obtener las combinaciones
    combinations = []
    # if 'combinations' in report_types:
    #     combinations = select_product_combinations()
    #     if not combinations:
    #         messagebox.showwarning("Advertencia", "No se seleccionaron combinaciones. No se generará reporte de combinaciones.")
    #         report_types.remove('combinations')
    #         if not report_types:
    #             return

     # 4. Obtener datos con la nueva consulta que incluye createdAt
  
    main_df = principalQuery()
            
    proccesdata('2021-01-01', '2022-12-11', report_types, combinations, main_df)
    proccesdata('2022-12-20', '2023-12-31', report_types, combinations, main_df)
    proccesdata('2024-01-01', '2025-01-31', report_types, combinations, main_df)

if __name__ == "__main__":
    main()
