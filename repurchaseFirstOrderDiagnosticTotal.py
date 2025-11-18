import pandas as pd
import json
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import simpledialog, messagebox, scrolledtext
from modules.database_queries import execute_query


# Diccionarios de mapeo (manteniendo los originales)
variables_dict = {
    1: "HEAD OF HAIR",
    2: "HAIR CONDITION",
    3: "TEXTURE",
    4: "HAIR TYPE",
    5: "HAIR STYLE",
    6: "AGE",
    7: "GLOVE SIZE",
    8: "NATURAL COLOR",
    9: "NATURAL SHADE",
    10: "CURRENT GRAY",
    11: "GRAY CONCENTRATION",
    12: "GRAY DIFFICULTY",
    13: "STUBBORN GRAYS",
    14: "QUIZ TYPE",
    15: "PRIMARY OBJECTIVE",
    16: "STAY THE SAME",
    17: "BEARD LENGTH",
    18: "BEARD THICK",
    19: "BEARD SCRUB",
    20: "FAMILY COLOR",
    21: "DESIRED SHADE",
    24: "DESIRED COVERAGE",
    27: "USER FIRST NAME",
    29: "USER EMAIL",
    30: "SALT & PEPPER LOOK",
    31: "VELOCITY",
    32: "COLOR INCREASE",
    33: "SHIPPER",
    34: "PURCHASE NUMBER",
    35: "USER LAST NAME",
    36: "EXPERIENCE WITH COLOR",
    37: "COLORED HAIR",
    39: "HAIR TREATMENT",
    40: "B - EXPERIENCE WITH COLOR",
    41: "SKIN REACTION",
    42: "CONTACT CUSTOMER SERVICE",
    43: "PURCHASE PLAN",
    44: "SUBSCRIPTION STARTING DATE",
    45: "USE DEVELOPER FROM ITALY"
}

values_dict = {
    4: "BALDING",
    5: "FULL",
    6: "STARTING TO THIN",
    7: "ULTRA DRY HAIR",
    8: "NORMAL",
    9: "DAMAGED HAIR",
    10: "THIN / FINE HAIR",
    11: "NORMAL HAIR",
    12: "THICK / COARSE",
    13: "CAUCASIAN",
    14: "AFRICAN",
    15: "ASIAN",
    16: "HS1",
    17: "HS2",
    18: "HS3",
    19: "HS4",
    20: "HS5",
    21: "HS6",
    22: "HS7",
    23: "HS8",
    24: "< 30",
    25: "30s",
    26: "40s",
    27: "50s",
    28: "60",
    29: "S",
    30: "M",
    31: "L",
    32: "XL",
    33: "BLACK",
    34: "DARK BROWN",
    35: "LIGHT BROWN",
    36: "RED",
    37: "DARK BLOND",
    38: "LIGHT BLOND",
    39: "JET BLACK",
    40: "BLACK",
    41: "SOFT BLACK",
    42: "DARK BROWN",
    43: "MEDIUM-DARK BROWN",
    44: "MEDIUM-LIGHT BROWN",
    45: "MEDIUM BROWN",
    46: "LIGHT BROWN",
    47: "DARK BLOND",
    48: "WARM-DARK BLOND",
    49: "COOL-MEDIUM BLOND",
    50: "WARM-MEDIUM BLOND",
    51: "COOL-LIGHT BLOND",
    52: "WARM-LIGHT BLOND",
    53: "LIGHTEST BLOND",
    54: "DARK AUBURN",
    55: "LIGHT AUBURN",
    56: "30% OR LESS",
    57: "30%-50%",
    58: "50%-75%",
    59: "NONE",
    60: "MORE THAN 75%",
    61: "SCATTERED",
    62: "CONCENTRATED",
    63: "EASY",
    64: "DIFFICULT",
    65: "DON'T KNOW",
    66: "YES",
    67: "HAIR",
    68: "BEARD",
    69: "NO",
    70: "COVER GRAY & MATCH SHADE",
    71: "CHANGE COLOR",
    72: "YES",
    73: "NO",
    74: "SHORT",
    75: "MEDIUM",
    76: "LONG",
    77: "FULL",
    78: "AVERAGE",
    79: "PATCHY",
    80: "MOISTURIZING",
    81: "ENERGIZING",
    82: "SENSITIVE",
    92: "FULL",
    93: "PARTIAL",
    94: "TARGETED / TOUCH UP",
    95: "FULL COVERAGE - IMMEDIATE",
    96: "FULL COVERAGE - GRADUAL",
    97: "SALT & PEPPER TARGETED",
    98: "SALT & PEPPER BLENDED",
    99: "30% OR LESS",
    100: "30%-50%",
    101: "50%-75%",
    102: "MORE THAN 75%",
    104: "INCREASE",
    105: "DECREASE",
    106: "SAME",
    107: "SMALL",
    108: "BIG",
    109: "1",
    110: "2",
    111: "3",
    112: "CURRENTLY DYED",
    113: "Highlights**",
    114: "Never colored",
    115: "Regular Hair Dye",
    116: "I've Colored",
    117: "COILY HAIR",
    118: "B - Currently Dyed",
    119: "B - I've colored",
    120: "B - Never colored",
    121: "YES",
    122: "NO",
    123: "CONTACT CUSTOMER SERVICE",
    124: "CONTINUE QUIZ",
    125: "OTO",
    126: "SUBSCRIPTION",
    127: "HAIR_INSTRUCTIONS",
    128: "BEARD_INSTRUCTIONS",
    129: "BOTH_INSTRUCTIONS",
    130: "NORMAL",
    131: "DRY",
    132: "YES",
    133: "NO"
}

# Nuevo diccionario para productos
product_dict = {
    'IT00000000000000000000000000000001': 'Big Shipper: Small Gloves',
    'IT00000000000000000000000000000002': '30ml Colorant - Medium Brown',
    'IT00000000000000000000000000000005': 'Small Shipper: X-Large Gloves',
    'IT00000000000000000000000000000006': '30ml Colorant - Jet-Black',
    'IT00000000000000000000000000000007': '30ml Colorant - Black',
    'IT00000000000000000000000000000008': '30ml Colorant - Soft-Black',
    'IT00000000000000000000000000000009': '30ml Colorant - Dark Brown',
    'IT00000000000000000000000000000010': '30ml Colorant - Medium-Dark Brown',
    'IT00000000000000000000000000000011': '30ml Colorant - Medium-Light Brown',
    'IT00000000000000000000000000000012': '30ml Colorant - Light Brown',
    'IT00000000000000000000000000000013': '30ml Colorant - Dark Blond',
    'IT00000000000000000000000000000014': '30ml Colorant - Warm- Dark Blond',
    'IT00000000000000000000000000000015': '30ml Colorant - Cool- Medium Blond',
    'IT00000000000000000000000000000016': '30ml Colorant - Warm- Medium Blond',
    'IT00000000000000000000000000000017': '30ml Colorant - Cool- Light Blond',
    'IT00000000000000000000000000000018': '30ml Colorant - Warm- Light Blond',
    'IT00000000000000000000000000000019': '30ml Colorant - Lightest Blond',
    'IT00000000000000000000000000000020': '30ml Colorant - Dark Auburn',
    'IT00000000000000000000000000000021': '30ml Colorant - Light Auburn',
    'IT00000000000000000000000000000022': '30ml Developer - 20 Vol',
    'IT00000000000000000000000000000023': '30ml Developer - 10 Vol',
    'IT00000000000000000000000000000024': 'Tray with Spatula',
    'IT00000000000000000000000000000025': '7 Min: Short Beard Brush',
    'IT00000000000000000000000000000026': '10 Min: Short Beard Brush',
    'IT00000000000000000000000000000028': '10 Min: Medium Beard Brush',
    'IT00000000000000000000000000000030': '10 Min: Long Beard Brush',
    'IT00000000000000000000000000000031': 'Small Shipper: Normal Hair Treatment',
    'IT00000000000000000000000000000032': 'Small Shipper: Thin Hair Treatment',
    'IT00000000000000000000000000000033': 'Small Shipper: Ultra Dry Hair Treatment',
    'IT00000000000000000000000000000034': 'Small Shipper: Damaged Hair Treatment',
    'IT00000000000000000000000000000035': 'Small Shipper: Coily Hair Treatment',
    'IT00000000000000000000000000000036': 'Small Shipper: Moisturizing',
    'IT00000000000000000000000000000037': 'Small Shipper: Sensitive',
    'IT00000000000000000000000000000038': 'Small Shipper: Energizing',
    'IT00000000000000000000000000000039': 'Small Shipper',
    'IT00000000000000000000000000000040': 'Big Shipper',
    'IT00000000000000000000000000000041': 'Small Shipper: Small Gloves',
    'IT00000000000000000000000000000042': 'Big Shipper: Medium Gloves',
    'IT00000000000000000000000000000043': 'Big Shipper: Large Gloves',
    'IT00000000000000000000000000000044': 'Big Shipper: X-Large Gloves',
    'IT00000000000000000000000000000045': '45ml Colorant - Jet-Black',
    'IT00000000000000000000000000000046': '45ml Colorant - Black',
    'IT00000000000000000000000000000047': '45ml Colorant - Soft-Black',
    'IT00000000000000000000000000000048': '45ml Colorant - Dark Brown',
    'IT00000000000000000000000000000049': '45ml Colorant - Medium-Dark Brown',
    'IT00000000000000000000000000000050': '45ml Colorant - Medium Brown',
    'IT00000000000000000000000000000051': '45ml Colorant - Medium-Light Brown',
    'IT00000000000000000000000000000052': '45ml Colorant: Light Brown',
    'IT00000000000000000000000000000053': '45ml Colorant - Dark Blond',
    'IT00000000000000000000000000000054': '45ml Colorant - Warm-Dark Blond',
    'IT00000000000000000000000000000055': '45ml Colorant - Cool-Medium Blond',
    'IT00000000000000000000000000000056': '45ml Colorant - Warm-Medium Blond',
    'IT00000000000000000000000000000057': '45ml Colorant - Cool - Light Blond',
    'IT00000000000000000000000000000058': '45ml Colorant - Warm - Light Blond',
    'IT00000000000000000000000000000059': '45ml Colorant - Lightest Blond',
    'IT00000000000000000000000000000060': '45ml Colorant - Dark Auburn',
    'IT00000000000000000000000000000061': '45ml Colorant - Light Auburn',
    'IT00000000000000000000000000000062': '45ml Developer - 30 Vol',
    'IT00000000000000000000000000000063': '45ml Developer - 20 Vol',
    'IT00000000000000000000000000000064': '5 Min: Big Shipper - Brush and Tray',
    'IT00000000000000000000000000000065': '7 Min: Big Shipper - Brush and Tray',
    'IT00000000000000000000000000000066': '10 Min: Big Shipper - Brush and Tray',
    'IT00000000000000000000000000000067': '5 Min: Big Shipper - Brush+Tray+Spatula',
    'IT00000000000000000000000000000068': '7 Min: Big Shipper - Brush+Tray+Spatula',
    'IT00000000000000000000000000000069': '10 Min: Big Shipper - Brush+Tray+Spatula',
    'IT00000000000000000000000000000070': 'Big Shipper: Normal Hair Treatment',
    'IT00000000000000000000000000000071': 'Big Shipper: Thin Hair Treatment',
    'IT00000000000000000000000000000072': 'Big Shipper: Ultra Dry Hair Treatment',
    'IT00000000000000000000000000000073': 'Big Shipper: Damaged Hair Treatment',
    'IT00000000000000000000000000000074': 'Big Shipper: 4C / Afro-Textured Hair',
    'IT00000000000000000000000000000075': 'Box 2 MIX Empty - Big Shipper',
    'IT00000000000000000000000000000076': 'Box 2 MIX Pre-Packed - Small Shipper',
    'IT00000000000000000000000000000004': 'Small Shipper: Large Gloves',
    'IT00000000000000000000000000000003': 'Small Shipper: Medium Gloves',
    'IT00000000000000000000000000000029': '7 Min: Long Beard Brush',
    'IT00000000000000000000000000000027': '7 Min: Medium Beard Brush',
    'IT00000000000000000000000000000103': 'Beard Scrub 7ml Sensitive',
    'IT00000000000000000000000000000104': 'Beard Scrub 7ml Energizing',
    'IT00000000000000000000000000000082': 'Lightest Blond Hair and Beard Dye',
    'IT00000000000000000000000000000083': 'Medium Blond Hair and Beard Dye',
    'IT00000000000000000000000000000084': 'Auburn Hair and Beard Dye',
    'IT00000000000000000000000000000085': 'Light Brown Hair and Beard Dye',
    'IT00000000000000000000000000000086': 'Medium Brown Hair and Beard Dye',
    'IT00000000000000000000000000000087': 'Dark Brown Hair and Beard Dye',
    'IT00000000000000000000000000000088': 'Natural Black Hair and Beard Dye',
    'IT00000000000000000000000000000089': 'Jet Black Hair and Beard Dye',
    'IT00000000000000000000000000000098': 'Dark Blond Hair and Beard Dye',
    'IT00000000000000000000000000000102': 'Black Hair and Beard Dye',
    'IT00000000000000000000000000000105': 'Jet Black Hair and Beard Dye FBM',
    'IT00000000000000000000000000000106': 'Dark Brown Hair and Beard Dye FBM',
    'IT00000000000000000000001004170006': 'Partial Coverage Beard Kit',
    'IT00000000000000000000001004170007': 'Full Coverage Beard Kit'
}

def parse_diagnostico(diagnostico_json):
    # (Mantener esta función igual)
    productos = []
    
    try:
        data = json.loads(diagnostico_json)
        values_data = data.get('values', {})
        
        # Manejar ambos formatos: objeto con keys numéricas o array
        if isinstance(values_data, dict):
            # Formato: {"396654": {"value": "14", "variable": "4"}}
            for key, item in values_data.items():
                variable = int(item.get('variable', 0))
                value = item.get('value', '')
                productos.append((variable, str(value)))
                
        elif isinstance(values_data, list):
            # Formato: [{"value": 13, "variable": 4}]
            for item in values_data:
                variable = int(item.get('variable', 0))
                value = item.get('value', '')
                productos.append((variable, str(value)))
                
    except (json.JSONDecodeError, TypeError, ValueError) as e:
        print(f"Error parsing JSON: {e}")
        return []
    
    return productos

def traducir_producto(variable, value):
    # (Mantener esta función igual)
    variable_trad = variables_dict.get(variable, f"VARIABLE_{variable}")
    
    # Intentar convertir value a número para buscar en values_dict
    try:
        value_num = int(value)
        value_trad = values_dict.get(value_num, str(value))
    except ValueError:
        value_trad = value
    
    return f"{variable_trad}:{value_trad}"

def parse_items(items_json):
    # (Mantener esta función igual)
    try:
        # Intentar parsear como JSON string primero
        if isinstance(items_json, str):
            items_list = json.loads(items_json)
        else:
            # Si ya es una lista (puede pasar dependiendo de cómo vengan los datos)
            items_list = items_json
        
        return items_list
        
    except (json.JSONDecodeError, TypeError, ValueError) as e:
        print(f"Error parsing items JSON: {e}")
        return []

def ajustar_ancho_columnas(archivo_excel):
    # (Mantener esta función igual)
    try:
        # Cargar el workbook
        wb = load_workbook(archivo_excel)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Ajustar el ancho de cada columna basado en el contenido
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            # Calcular la longitud del contenido
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Ajustar el ancho de la columna (agregar un poco de padding)
                adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres de ancho
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar los cambios
        wb.save(archivo_excel)
        
    except Exception as e:
        print(f"Error ajustando el ancho de columnas: {e}")

def procesar_datos_diagnostico(df, filtro_items=None):
    """
    Procesa el DataFrame y genera el análisis de recompra por producto del diagnóstico
    Si se proporciona filtro_items, solo incluye órdenes que contengan ese item
    """
    # Filtrar el DataFrame si se especifica un item
    if filtro_items:
        df_filtrado = df[df['first_order_items'].apply(
            lambda x: filtro_items in parse_items(x) if pd.notna(x) else False
        )]
    else:
        df_filtrado = df
    
    # Diccionario para almacenar resultados
    resultados = defaultdict(lambda: {'recompraron': 0, 'no_recompraron': 0})
    
    if df_filtrado.size == 0:
        return pd.DataFrame()
    
    for _, row in df_filtrado.iterrows():
        diagnostico = row['diagnostic']
        total_compras = row['total_custom_orders']
        
        # Saltar si no hay diagnóstico
        if pd.isna(diagnostico):
            continue
            
        # Parsear el diagnóstico
        productos_raw = parse_diagnostico(diagnostico)
        
        # Procesar cada producto del diagnóstico
        for variable, value in productos_raw:
            # Excluir variables relacionadas con USER FIRST NAME (variable 27)
            if variable == 27:
                continue
                
            producto = traducir_producto(variable, value)
            
            # Contar recompra o no recompra
            if total_compras > 1:
                resultados[producto]['recompraron'] += 1
            else:
                resultados[producto]['no_recompraron'] += 1
    
    # Convertir a DataFrame
    df_resultado = pd.DataFrame([
        {
            'Producto': producto,
            'Usuarios que recompraron': stats['recompraron'],
            'Usuarios que no recompraron': stats['no_recompraron'],
            'Total de usuarios': stats['recompraron'] + stats['no_recompraron'],
            'Porcentaje de recompra': (stats['recompraron'] / (stats['recompraron'] + stats['no_recompraron'])) * 100 
            if (stats['recompraron'] + stats['no_recompraron']) > 0 else 0
        }
        for producto, stats in resultados.items()
    ])
    
    # Ordenar por total de usuarios (descendente)
    df_resultado = df_resultado.sort_values('Total de usuarios', ascending=False)
    
    return df_resultado

def procesar_datos_combinaciones_diagnostico(df, combinaciones):
    """
    Procesa el DataFrame y genera el análisis de recompra por combinaciones de variables:values
    """
    # Diccionario para almacenar resultados
    resultados = defaultdict(lambda: {'recompraron': 0, 'no_recompraron': 0})
    
    if df.size == 0:
        return pd.DataFrame()
    
    for _, row in df.iterrows():
        diagnostico = row['diagnostic']
        total_compras = row['total_custom_orders']
        
        # Saltar si no hay diagnóstico
        if pd.isna(diagnostico):
            continue
            
        # Parsear el diagnóstico
        productos_raw = parse_diagnostico(diagnostico)
        diagnostico_traducido = [traducir_producto(variable, value) for variable, value in productos_raw]
        
        # Verificar cada combinación
        for nombre_combinacion, variables_combinacion in combinaciones.items():
            # Verificar si todos los elementos de la combinación están en el diagnóstico
            todos_presentes = all(variable in diagnostico_traducido for variable in variables_combinacion)
            
            if todos_presentes:
                # Contar recompra o no recompra para esta combinación
                if total_compras > 1:
                    resultados[nombre_combinacion]['recompraron'] += 1
                else:
                    resultados[nombre_combinacion]['no_recompraron'] += 1
    
    # Convertir a DataFrame
    df_resultado = pd.DataFrame([
        {
            'Combinación': combinacion,
            'Variables': ', '.join(combinaciones[combinacion]),
            'Usuarios que recompraron': stats['recompraron'],
            'Usuarios que no recompraron': stats['no_recompraron'],
            'Total de usuarios': stats['recompraron'] + stats['no_recompraron'],
            'Porcentaje de recompra': (stats['recompraron'] / (stats['recompraron'] + stats['no_recompraron'])) * 100 
            if (stats['recompraron'] + stats['no_recompraron']) > 0 else 0
        }
        for combinacion, stats in resultados.items()
    ])
    
    # Ordenar por total de usuarios (descendente)
    df_resultado = df_resultado.sort_values('Total de usuarios', ascending=False)
    
    return df_resultado

def procesar_rango_fechas(start_date, end_date, nombre_rango):
    """
    Procesa un rango de fechas y devuelve los DataFrames para diagnóstico
    """
    query = f"""
        WITH first_orders AS (
        SELECT fo.id AS order_id, fo.customer_id, fo.created_at
        FROM bi.fact_orders fo
        WHERE fo.is_first_order = 1
            AND fo.order_plan = 'OTO'
            AND fo.status NOT IN ('CANCELLED','PAYMENT_ERROR')
            AND fo.created_at BETWEEN '{start_date}' AND '{end_date}'
            AND EXISTS (
            SELECT 1
            FROM bi.fact_sales_order_items soi
            WHERE soi.salesOrderId = fo.id
                AND soi.category = 'IG00000000000000000000000000000029' 
            )
        ),
        custom_orders_count AS (
        SELECT fo.customer_id, COUNT(DISTINCT fo.id) AS total_custom_orders
        FROM bi.fact_orders fo
        WHERE fo.status NOT IN ('CANCELLED','PAYMENT_ERROR')
            AND EXISTS (
            SELECT 1
            FROM bi.fact_sales_order_items soi
            WHERE soi.salesOrderId = fo.id
                AND soi.category IN (
                'IG00000000000000000000000000000029',
                'IG00000000000000000000000000000028'
                )
            )
        GROUP BY fo.customer_id
        ),
        first_order_items AS (
        SELECT soi.salesOrderId AS order_id,
                JSON_ARRAYAGG(soi.itemId) AS item_list
        FROM bi.fact_sales_order_items soi
        WHERE soi.salesOrderId IN (SELECT order_id FROM first_orders)
        GROUP BY soi.salesOrderId
        )
        SELECT
        fo.customer_id,
        (
            SELECT psoi.additionalFields->>"$.diagnostic"
            FROM prod_sales_and_subscriptions.sales_order_items psoi
            WHERE psoi.salesOrderId = fo.order_id
            AND psoi.itemId IN ('IT00000000000000000000001004170007','IT00000000000000000000001004170006')
            ORDER BY psoi.itemId
            LIMIT 1
        ) AS diagnostic,
        coc.total_custom_orders,
        foi.item_list AS first_order_items,
        fo.created_at AS first_order_date
        FROM first_orders fo
        JOIN custom_orders_count coc ON coc.customer_id = fo.customer_id
        JOIN first_order_items foi   ON foi.order_id     = fo.order_id;
    """
    
    df = execute_query(query)
    
    # Procesar los datos de diagnóstico para las 3 páginas
    df_diagnostico_total = procesar_datos_diagnostico(df)
    df_diagnostico_item22 = procesar_datos_diagnostico(df, 'IT00000000000000000000000000000022')
    df_diagnostico_item23 = procesar_datos_diagnostico(df, 'IT00000000000000000000000000000023')
    
    # Añadir nombre del rango a los DataFrames
    df_diagnostico_total['Rango'] = nombre_rango
    df_diagnostico_item22['Rango'] = nombre_rango
    df_diagnostico_item23['Rango'] = nombre_rango
    
    return {
        'diagnostico_total': df_diagnostico_total,
        'diagnostico_item22': df_diagnostico_item22,
        'diagnostico_item23': df_diagnostico_item23,
        'datos_completos': df  # Devolvemos también los datos completos para las combinaciones
    }

def combinar_dataframes(lista_dataframes, nombres_rangos, tipo_analisis="diagnostico"):
    """
    Combina múltiples DataFrames en uno solo con el formato de columnas solicitado
    """
    if not lista_dataframes:
        return pd.DataFrame()
    
    # Mapeo de códigos a nombres completos
    codigos_nombres = {
        'CTF': 'Control Timeframe',
        'DVL': 'OTO Developer change',
        'AMZ': 'Post Amazon Launch'
    }
    
    # Calcular el máximo de "Total de usuarios" por cada rango
    maximos_por_rango = {}
    for i, df in enumerate(lista_dataframes):
        if df is not None and not df.empty and 'Total de usuarios' in df.columns:
            codigo_rango = list(nombres_rangos.keys())[i]
            maximos_por_rango[codigo_rango] = df['Total de usuarios'].max()
        else:
            maximos_por_rango[list(nombres_rangos.keys())[i]] = 0
    
    # Calcular totales por variable para cada rango
    totales_por_variable_rango = {}
    for i, df in enumerate(lista_dataframes):
        if df is None or df.empty:
            continue
            
        codigo_rango = list(nombres_rangos.keys())[i]
        totales_por_variable_rango[codigo_rango] = {}
        
        # Extraer variable de cada producto
        if tipo_analisis == "diagnostico":
            df['variable_only'] = df['Producto'].apply(lambda x: x.split(':')[0] if ':' in x else '')
        else:
            # Para combinaciones, usamos la variable directamente
            df['variable_only'] = df['Combinación']
        
        # Calcular total por variable
        for variable in df['variable_only'].unique():
            if variable:
                total_variable = df[df['variable_only'] == variable]['Total de usuarios'].sum()
                totales_por_variable_rango[codigo_rango][variable] = total_variable
    
    # Crear un diccionario para almacenar los datos combinados
    datos_combinados = {}
    
    # Procesar cada DataFrame
    for i, df in enumerate(lista_dataframes):
        if df is None or df.empty:
            continue
            
        # Determinar el código y nombre del rango
        codigo_rango = list(nombres_rangos.keys())[i]
        nombre_rango = nombres_rangos[codigo_rango]
        maximo_rango = maximos_por_rango[codigo_rango]
        
        # Extraer variable de cada producto para este DataFrame
        if tipo_analisis == "diagnostico":
            df['variable_only'] = df['Producto'].apply(lambda x: x.split(':')[0] if ':' in x else '')
        else:
            df['variable_only'] = df['Combinación']
        
        for _, row in df.iterrows():
            if tipo_analisis == "diagnostico":
                clave = row['Producto']
                variable = clave.split(':')[0]
                value = clave.split(':')[1] if ':' in clave else ''
            else:  # combinaciones
                clave = row['Combinación']
                variable = clave
                value = row['Variables']
            
            if clave not in datos_combinados:
                datos_combinados[clave] = {
                    'Variable': variable,
                    'Value': value,
                }
                
                # Inicializar todas las columnas para los 3 rangos
                for codigo in nombres_rangos.keys():
                    datos_combinados[clave][f'Usuarios que recompraron {codigo}'] = 0
                    datos_combinados[clave][f'Usuarios que no recompraron {codigo}'] = 0
                    datos_combinados[clave][f'Total de usuarios {codigo}'] = 0
                    datos_combinados[clave][f'% del máximo {codigo}'] = 0
                    datos_combinados[clave][f'% por variable {codigo}'] = 0  # Nueva columna
                    datos_combinados[clave][f'{nombres_rangos[codigo]}'] = 0
            
            # Asignar los valores para este rango
            datos_combinados[clave][f'Usuarios que recompraron {codigo_rango}'] = row['Usuarios que recompraron']
            datos_combinados[clave][f'Usuarios que no recompraron {codigo_rango}'] = row['Usuarios que no recompraron']
            datos_combinados[clave][f'Total de usuarios {codigo_rango}'] = row['Total de usuarios']
            
            # Calcular el porcentaje basado en el valor máximo
            if maximo_rango > 0:
                porcentaje_maximo = (row['Total de usuarios'] / maximo_rango) * 100
            else:
                porcentaje_maximo = 0
            datos_combinados[clave][f'% del máximo {codigo_rango}'] = porcentaje_maximo
            
            # Calcular el porcentaje por variable
            total_variable = totales_por_variable_rango[codigo_rango].get(variable, 0)
            if total_variable > 0:
                porcentaje_variable = (row['Total de usuarios'] / total_variable) * 100
            else:
                porcentaje_variable = 0
            datos_combinados[clave][f'% por variable {codigo_rango}'] = porcentaje_variable
            
            datos_combinados[clave][f'{nombre_rango}'] = row['Porcentaje de recompra']
    
    # Convertir a DataFrame
    df_combinado = pd.DataFrame(list(datos_combinados.values()))
    
    # Reordenar las columnas
    column_order = ['Variable', 'Value']
    for codigo in nombres_rangos.keys():
        column_order.extend([
            f'Usuarios que recompraron {codigo}',
            f'Usuarios que no recompraron {codigo}',
            f'Total de usuarios {codigo}',
            f'% del máximo {codigo}',
            f'% por variable {codigo}',  # Nueva columna
            f'{nombres_rangos[codigo]}'
        ])
    
    # Asegurarse de que todas las columnas existan en el DataFrame
    existing_columns = [col for col in column_order if col in df_combinado.columns]
    df_combinado = df_combinado[existing_columns]
    
    # Ordenar por Variable y Value (alfabéticamente)
    df_combinado = df_combinado.sort_values(['Variable', 'Value'])
    
    # Formatear las columnas de porcentaje para incluir el símbolo %
    for codigo in nombres_rangos.keys():
        for porcentaje_col in [f'% del máximo {codigo}', f'% por variable {codigo}']:
            if porcentaje_col in df_combinado.columns:
                df_combinado[porcentaje_col] = df_combinado[porcentaje_col].apply(
                    lambda x: f"{x:.2f}%" if pd.notna(x) and isinstance(x, (int, float)) else x
                )
        
        recompra_col = f'{nombres_rangos[codigo]}'
        if recompra_col in df_combinado.columns:
            df_combinado[recompra_col] = df_combinado[recompra_col].apply(
                lambda x: f"{x:.2f}%" if pd.notna(x) and isinstance(x, (int, float)) else x
            )
    
    return df_combinado

def main():
    # Preguntar qué tipo de análisis realizar
    root = tk.Tk()
    root.withdraw()  # Ocultar la ventana principal
    
    opcion = messagebox.askquestion("Tipo de Análisis", 
                                   "¿Desea realizar análisis por combinaciones?\n\n" +
                                   "Sí: Análisis por combinaciones de diagnóstico\n" +
                                   "No: Análisis tradicional por diagnóstico")
    
    realizar_combinaciones = (opcion == 'yes')
    
    # Definir las combinaciones preestablecidas
    combinaciones_predefinidas = {
        "Exp. Color: Currently Dyed + Skin Reaction: NO": ["B - EXPERIENCE WITH COLOR:B - Currently Dyed", "SKIN REACTION:NO"],
        "Exp. Color: I've colored + Skin Reaction: NO": ["B - EXPERIENCE WITH COLOR:B - I've colored", "SKIN REACTION:NO"],
        "Exp. Color: Never colored + Skin Reaction: NO": ["B - EXPERIENCE WITH COLOR:B - Never colored", "SKIN REACTION:NO"],
        "Exp. Color: Currently Dyed + Skin Reaction: YES": ["B - EXPERIENCE WITH COLOR:B - Currently Dyed", "SKIN REACTION:YES"],
        "Exp. Color: I've colored + Skin Reaction: YES": ["B - EXPERIENCE WITH COLOR:B - I've colored", "SKIN REACTION:YES"],
        "Exp. Color: Never colored + Skin Reaction: YES": ["B - EXPERIENCE WITH COLOR:B - Never colored", "SKIN REACTION:YES"]
    }
    
    # Definir los rangos de fechas y sus códigos/nombres
    rangos_fechas = [
        # {'start': '2021-01-01', 'end': '2022-12-11', 'codigo': 'CTF', 'nombre': 'Control Timeframe'},
        # {'start': '2022-12-20', 'end': '2023-12-31', 'codigo': 'DVL', 'nombre': 'OTO Developer change'},
        {'start': '2025-01-01', 'end': '2025-04-01', 'codigo': 'Q1 - 2025', 'nombre': 'Q1 - 2025'}
    ]
    
    # Crear diccionario de nombres de rangos
    nombres_rangos = {rango['codigo']: rango['nombre'] for rango in rangos_fechas}
    
    # Procesar cada rango de fechas
    resultados = []
    datos_completos = []
    for rango in rangos_fechas:
        print(f"Procesando rango {rango['nombre']}: {rango['start']} a {rango['end']}")
        try:
            resultado = procesar_rango_fechas(rango['start'], rango['end'], rango['codigo'])
            resultados.append(resultado)
            datos_completos.append(resultado['datos_completos'])
        except Exception as e:
            print(f"Error procesando rango {rango['nombre']}: {e}")
            # Agregar DataFrames vacíos para mantener la estructura
            resultados.append({
                'diagnostico_total': pd.DataFrame(),
                'diagnostico_item22': pd.DataFrame(),
                'diagnostico_item23': pd.DataFrame(),
                'datos_completos': pd.DataFrame()
            })
            datos_completos.append(pd.DataFrame())
    
    # Guardar en un solo archivo Excel
    nombre_archivo = "analisis_recompra_consolidado.xlsx"
    
    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        if realizar_combinaciones:
            # Procesar combinaciones para cada rango
            combinaciones_resultados = []
            for i, df in enumerate(datos_completos):
                if df.empty:
                    combinaciones_resultados.append(pd.DataFrame())
                else:
                    resultado_combinaciones = procesar_datos_combinaciones_diagnostico(df, combinaciones_predefinidas)
                    resultado_combinaciones['Rango'] = list(nombres_rangos.keys())[i]
                    combinaciones_resultados.append(resultado_combinaciones)
            
            # Combinar resultados de combinaciones
            df_combinaciones_combinado = combinar_dataframes(combinaciones_resultados, nombres_rangos, "combinaciones")
            
            if not df_combinaciones_combinado.empty:
                df_combinaciones_combinado.to_excel(writer, sheet_name='Combinaciones', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='Combinaciones', index=False)
        else:
            # Procesar diagnóstico tradicional
            diagnosticos_total = [r['diagnostico_total'] for r in resultados]
            diagnosticos_item22 = [r['diagnostico_item22'] for r in resultados]
            diagnosticos_item23 = [r['diagnostico_item23'] for r in resultados]
            
            df_diagnostico_total_combinado = combinar_dataframes(diagnosticos_total, nombres_rangos)
            df_diagnostico_item22_combinado = combinar_dataframes(diagnosticos_item22, nombres_rangos)
            df_diagnostico_item23_combinado = combinar_dataframes(diagnosticos_item23, nombres_rangos)
            
            if not df_diagnostico_total_combinado.empty:
                df_diagnostico_total_combinado.to_excel(writer, sheet_name='Todos los diagnósticos', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='Todos los diagnósticos', index=False)
                
            if not df_diagnostico_item22_combinado.empty:
                df_diagnostico_item22_combinado.to_excel(writer, sheet_name='Con Developer 20Vol', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='Con Developer 20Vol', index=False)
                
            if not df_diagnostico_item23_combinado.empty:
                df_diagnostico_item23_combinado.to_excel(writer, sheet_name='Con Developer 10Vol', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='Con Developer 10Vol', index=False)
    
    # Ajustar el ancho de las columnas
    ajustar_ancho_columnas(nombre_archivo)
    
    print(f"Análisis completado. Resultados consolidados guardados en: {nombre_archivo}")

if __name__ == "__main__":
    main()