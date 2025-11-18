# -*- coding: utf-8 -*-
"""
Script final: Actualiza 'analisis_repurchase_cancelaciones.xlsx' desde los archivos de recompra y cancelaciones.
Autor: Juan Diego Bogotá (Cleverman)
Versión: Octubre 2025
-------------------------------------------------------
Funciones:
 ✅ Llena Recompra
 ✅ Llena Cancelaciones_Etnias (Total, Caucasian, African, Asian)
 ✅ Llena Cancelaciones_shade (Total + Shades específicos)
-------------------------------------------------------
Configuración:
 - Cambia los índices COL_IDX_* para elegir la columna de destino.
 - No modifica fórmulas ni hace cálculos.
"""

import pandas as pd
from openpyxl import load_workbook

# ================= CONFIGURACIÓN =================
RECOMPRA_FILE   = "analisis_recompra_consolidado.xlsx"
CANCEL_FILE     = "analisis_cancelaciones_2025-01-01_to_2025-04-01.xlsx"
MAESTRO_FILE    = "analisis_repurchase_cancelaciones.xlsx"

COL_IDX_RECOMPRA        = 3   # ej. "Q3 - 2025" → columna C
COL_IDX_CANCEL_ETNIAS   = 9
COL_IDX_CANCEL_SHADES   = 9

# ================= FUNCIONES AUXILIARES =================
def clean_percent(val):
    if isinstance(val, str):
        val = val.replace('%','').replace(',', '.').strip()
    try:
        return float(val)
    except:
        return None

def find_blocks_in_master(ws, razon_header_text="razon_procesada"):
    """Detecta bloques de tablas separadas por filas en blanco."""
    headers = {cell.value: i for i, cell in enumerate(ws[1], start=1)}
    col_razon = None
    for k,v in headers.items():
        if isinstance(k,str) and razon_header_text in k.lower():
            col_razon=v; break
    if not col_razon: return []
    blocks=[]
    r=2; maxr=ws.max_row
    while r<=maxr:
        while r<=maxr and ws.cell(row=r,column=col_razon).value in (None,""):
            r+=1
        if r>maxr: break
        start=r
        while r<=maxr and ws.cell(row=r,column=col_razon).value not in (None,""):
            r+=1
        end=r-1
        title=ws.cell(row=start,column=1).value or "Total"
        title=str(title).replace(" razon_procesada","").strip()
        if title.lower()=="razon_procesada": title="Total"
        blocks.append((start,end,title))
    return blocks

def _norm_shade(s:str)->str:
    """Normaliza nombres de shades."""
    return s.lower().replace(" ","").replace("-","")

def find_exact_shade_column(df, wanted_shade):
    """Match exacto de shade evitando confusiones ('Black' ≠ 'Soft-Black')."""
    if wanted_shade=="Total":
        return "porcentaje_cancelaciones"
    wanted=_norm_shade(wanted_shade)
    for c in df.columns:
        if isinstance(c,str) and c.lower().startswith("porcentaje_"):
            suffix=c.split(" - ",1)[-1].strip()
            if _norm_shade(suffix)==wanted:
                return c
    return None

# ================= LECTURA FUENTES =================
rep_df = pd.read_excel(RECOMPRA_FILE, sheet_name="Todos los diagnósticos")

etnias_src = pd.read_excel(CANCEL_FILE, sheet_name="Por Razon (Etnias)")
ETNIAS_COL_MAP = {
    "Total": "porcentaje_cancelaciones",
    "Caucasian": "porcentaje_caucasian",
    "African": "porcentaje_african",
    "Asian": "porcentaje_asian",
}

shades_src_all = pd.read_excel(CANCEL_FILE, sheet_name="Por Razon (Shades)")
shades_src = shades_src_all.iloc[0:15].copy()  # sólo filas 1–15

# ================= ABRIR MAESTRO =================
wb = load_workbook(MAESTRO_FILE)
summary_counts = {"Recompra":0, "Etnias":{}, "Shades":{}}

# ----------- RECOMPRA -----------
if "Recompra" in wb.sheetnames:
    ws=wb["Recompra"]
    headers={c.value:i for i,c in enumerate(ws[1],start=1)}
    col_var, col_val = headers.get("Variable"), headers.get("Value")
    rep_cols=[c for c in rep_df.columns if isinstance(c,str) and ('%' in c or 'Q' in c)]
    col_src=rep_cols[-1] if rep_cols else None
    if col_var and col_val and col_src:
        for r in range(2,ws.max_row+1):
            var=ws.cell(row=r,column=col_var).value
            val=ws.cell(row=r,column=col_val).value
            if not var or not val: continue
            m=rep_df[
                (rep_df["Variable"].astype(str).str.strip()==str(var).strip()) &
                (rep_df["Value"].astype(str).str.strip()==str(val).strip())
            ]
            if not m.empty:
                ws.cell(row=r,column=COL_IDX_RECOMPRA).value=clean_percent(m[col_src].values[0])
                summary_counts["Recompra"]+=1

# ----------- CANCELACIONES ETNIAS -----------
if "Cancelaciones_Etnias" in wb.sheetnames:
    ws=wb["Cancelaciones_Etnias"]
    headers={c.value:i for i,c in enumerate(ws[1],start=1)}
    col_razon=headers.get("razon_procesada")
    blocks=find_blocks_in_master(ws)
    grupos=["Total","Caucasian","African","Asian"]
    for (start,end,title),grupo in zip(blocks,grupos):
        col_name=ETNIAS_COL_MAP.get(grupo)
        if not col_name: continue
        count=0
        for r in range(start+1,end+1):
            razon=ws.cell(row=r,column=col_razon).value
            if not razon: continue
            m=etnias_src[etnias_src["razon_procesada"].astype(str).str.strip()==str(razon).strip()]
            if not m.empty and col_name in m.columns:
                ws.cell(row=r,column=COL_IDX_CANCEL_ETNIAS).value=clean_percent(m[col_name].values[0])
                count+=1
        summary_counts["Etnias"][grupo]=count

# ----------- CANCELACIONES SHADES -----------
if "Cancelaciones_shade" in wb.sheetnames:
    ws=wb["Cancelaciones_shade"]
    headers={c.value:i for i,c in enumerate(ws[1],start=1)}
    col_razon=headers.get("razon_procesada")
    blocks=find_blocks_in_master(ws)
    for (start,end,title) in blocks:
        col_name=find_exact_shade_column(shades_src,title)
        if not col_name: continue
        count=0
        for r in range(start+1,end+1):
            razon=ws.cell(row=r,column=col_razon).value
            if not razon: continue
            m=shades_src[shades_src["razon_procesada"].astype(str).str.strip()==str(razon).strip()]
            if not m.empty and col_name in m.columns:
                ws.cell(row=r,column=COL_IDX_CANCEL_SHADES).value=clean_percent(m[col_name].values[0])
                count+=1
        summary_counts["Shades"][title]=count

# ================= GUARDAR =================
wb.save(MAESTRO_FILE)

# ================= RESUMEN =================
print("✅ Archivo maestro actualizado correctamente.\n")
print(f"Recompra: {summary_counts['Recompra']} valores escritos.")
print("Etnias:")
for k,v in summary_counts["Etnias"].items():
    print(f"  {k}: {v}")
print("Shades:")
for k,v in summary_counts["Shades"].items():
    print(f"  {k}: {v}")
