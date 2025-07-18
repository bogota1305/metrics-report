import pandas as pd
import tkinter as tk
from tkinter import messagebox, Label, Button
from tkcalendar import Calendar
from modules.database_queries import execute_query
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

def open_rebuy_date_selector():
    """Interfaz para seleccionar fechas de análisis"""
    start_date = None  # Definición inicial
    end_date = None    # Definición inicial
    
    def get_dates():
        nonlocal start_date, end_date  # Ahora referencia variables existentes
        start_date = cal_start.get_date()
        end_date = cal_end.get_date()
        
        if not start_date or not end_date:
            messagebox.showerror("Error", "Selecciona ambas fechas.")
            return
            
        # Convertir a formato YYYY-MM-DD
        start_date = pd.to_datetime(start_date).strftime('%Y-%m-%d')
        end_date = pd.to_datetime(end_date).strftime('%Y-%m-%d')
        
        root.quit()
        root.destroy()

    # Configuración de la ventana
    root = tk.Tk()
    root.title("Selección de Rango de Fechas")
    
    Label(root, text="Fecha de inicio:").pack(pady=5)
    cal_start = Calendar(root, selectmode='day', date_pattern='yyyy-mm-dd')
    cal_start.pack(pady=5)

    Label(root, text="Fecha de fin:").pack(pady=5)
    cal_end = Calendar(root, selectmode='day', date_pattern='yyyy-mm-dd')
    cal_end.pack(pady=5)

    Button(root, text="Generar Reporte", command=get_dates).pack(pady=20)
    
    root.mainloop()
    return start_date, end_date

def get_product_classification():
    """Clasificación manual completa de productos basada en IDs"""
    # Primero definimos un diccionario de IDs a nombres
    id_to_name = {
        'IT00000000000000000000000000000001': 'Big Shipper: Small Gloves',
        'IT00000000000000000000000000000002': '30ml Colorant - Medium Brown',
        'IT00000000000000000000000000000005': 'Small Shipper: X-Large Gloves',
        'IT00000000000000000000000000000006': '30ml Colorant - Jet-Black',
        'IT00000000000000000000000000000007': '30ml Colorant - Black',
        'IT00000000000000000000000000000008': '30ml Colorant - Soft-Black',
        'IT00000000000000000000000000000009': '30ml Colorant - Dark Brown',
        'IT00000000000000000000000000000010': '30ml Colorant - Medium-Dark Brown',
        'IT00000000000000000000000000000011': '30ml Colorant - Medium-Light Brown',
        'IT00000000000000000000000000000012': '30ml Colorant - Light Brown',
        'IT00000000000000000000000000000013': '30ml Colorant - Dark Blond',
        'IT00000000000000000000000000000014': '30ml Colorant - Warm- Dark Blond',
        'IT00000000000000000000000000000015': '30ml Colorant - Cool- Medium Blond',
        'IT00000000000000000000000000000016': '30ml Colorant - Warm- Medium Blond',
        'IT00000000000000000000000000000017': '30ml Colorant - Cool- Light Blond',
        'IT00000000000000000000000000000018': '30ml Colorant - Warm- Light Blond',
        'IT00000000000000000000000000000019': '30ml Colorant - Lightest Blond',
        'IT00000000000000000000000000000020': '30ml Colorant - Dark Auburn',
        'IT00000000000000000000000000000021': '30ml Colorant - Light Auburn',
        'IT00000000000000000000000000000022': '30ml Developer - 20 Vol',
        'IT00000000000000000000000000000023': '30ml Developer - 10 Vol',
        'IT00000000000000000000000000000024': 'Tray with Spatula',
        'IT00000000000000000000000000000025': '7 Min: Short Beard Brush',
        'IT00000000000000000000000000000026': '10 Min: Short Beard Brush',
        'IT00000000000000000000000000000028': '10 Min: Medium Beard Brush',
        'IT00000000000000000000000000000030': '10 Min: Long Beard Brush',
        'IT00000000000000000000000000000031': 'Small Shipper: Normal Hair Treatment',
        'IT00000000000000000000000000000032': 'Small Shipper: Thin Hair Treatment',
        'IT00000000000000000000000000000033': 'Small Shipper: Ultra Dry Hair Treatment',
        'IT00000000000000000000000000000034': 'Small Shipper: Damaged Hair Treatment',
        'IT00000000000000000000000000000035': 'Small Shipper: Coily Hair Treatment',
        'IT00000000000000000000000000000036': 'Small Shipper: Moisturizing',
        'IT00000000000000000000000000000037': 'Small Shipper: Sensitive',
        'IT00000000000000000000000000000038': 'Small Shipper: Energizing',
        'IT00000000000000000000000000000039': 'Small Shipper',
        'IT00000000000000000000000000000040': 'Big Shipper',
        'IT00000000000000000000000000000041': 'Small Shipper: Small Gloves',
        'IT00000000000000000000000000000042': 'Big Shipper: Medium Gloves',
        'IT00000000000000000000000000000043': 'Big Shipper: Large Gloves',
        'IT00000000000000000000000000000044': 'Big Shipper: X-Large Gloves',
        'IT00000000000000000000000000000045': '45ml Colorant - Jet-Black',
        'IT00000000000000000000000000000046': '45ml Colorant - Black',
        'IT00000000000000000000000000000047': '45ml Colorant - Soft-Black',
        'IT00000000000000000000000000000048': '45ml Colorant - Dark Brown',
        'IT00000000000000000000000000000049': '45ml Colorant - Medium-Dark Brown',
        'IT00000000000000000000000000000050': '45ml Colorant - Medium Brown',
        'IT00000000000000000000000000000051': '45ml Colorant - Medium-Light Brown',
        'IT00000000000000000000000000000052': '45ml Colorant: Light Brown',
        'IT00000000000000000000000000000053': '45ml Colorant - Dark Blond',
        'IT00000000000000000000000000000054': '45ml Colorant - Warm-Dark Blond',
        'IT00000000000000000000000000000055': '45ml Colorant - Cool-Medium Blond',
        'IT00000000000000000000000000000056': '45ml Colorant - Warm-Medium Blond',
        'IT00000000000000000000000000000057': '45ml Colorant - Cool - Light Blond',
        'IT00000000000000000000000000000058': '45ml Colorant - Warm - Light Blond',
        'IT00000000000000000000000000000059': '45ml Colorant - Lightest Blond',
        'IT00000000000000000000000000000060': '45ml Colorant - Dark Auburn',
        'IT00000000000000000000000000000061': '45ml Colorant - Light Auburn',
        'IT00000000000000000000000000000062': '45ml Developer - 30 Vol',
        'IT00000000000000000000000000000063': '45ml Developer - 20 Vol',
        'IT00000000000000000000000000000064': '5 Min: Big Shipper - Brush and Tray',
        'IT00000000000000000000000000000065': '7 Min: Big Shipper - Brush and Tray',
        'IT00000000000000000000000000000066': '10 Min: Big Shipper - Brush and Tray',
        'IT00000000000000000000000000000067': '5 Min: Big Shipper - Brush+Tray+Spatula',
        'IT00000000000000000000000000000068': '7 Min: Big Shipper - Brush+Tray+Spatula',
        'IT00000000000000000000000000000069': '10 Min: Big Shipper - Brush+Tray+Spatula',
        'IT00000000000000000000000000000070': 'Big Shipper: Normal Hair Treatment',
        'IT00000000000000000000000000000071': 'Big Shipper: Thin Hair Treatment',
        'IT00000000000000000000000000000072': 'Big Shipper: Ultra Dry Hair Treatment',
        'IT00000000000000000000000000000073': 'Big Shipper: Damaged Hair Treatment',
        'IT00000000000000000000000000000074': 'Big Shipper: 4C / Afro-Textured Hair',
        'IT00000000000000000000000000000075': 'Box 2 MIX Empty - Big Shipper',
        'IT00000000000000000000000000000076': 'Box 2 MIX Pre-Packed - Small Shipper',
        'IT00000000000000000000000000000079': 'Gift Certificate',
        'IT00000000000000000000000000000080': 'Jaxon Lane Sampling',
        'IT00000000000000000000000000000004': 'Small Shipper: Large Gloves',
        'IT00000000000000000000000000000003': 'Small Shipper: Medium Gloves',
        'IT00000000000000000000000000000029': '7 Min: Long Beard Brush',
        'IT00000000000000000000000000000027': '7 Min: Medium Beard Brush',
        'IT00000000000000000000001004170001': 'Full Coverage Immediate Kit',
        'IT00000000000000000000001004170002': 'Salt & Pepper Targeted Kit',
        'IT00000000000000000000001004170003': 'Full Coverage Gradual Kit',
        'IT00000000000000000000001004170004': 'Salt & Pepper Blended Kit',
        'IT00000000000000000000001004170005': 'Targeted / Touch Up Kit',
        'IT00000000000000000000001004170006': 'Partial Coverage Beard Kit',
        'IT00000000000000000000001004170007': 'Full Coverage Beard Kit',
        'IT00000000000000000000000000000081': 'PROMO Tweezers',
        'IT00000000000000000000000000000082': 'Lightest Blond Hair and Beard Dye',
        'IT00000000000000000000000000000083': 'Medium Blond Hair and Beard Dye',
        'IT00000000000000000000000000000084': 'Auburn Hair and Beard Dye',
        'IT00000000000000000000000000000085': 'Light Brown Hair and Beard Dye',
        'IT00000000000000000000000000000086': 'Medium Brown Hair and Beard Dye',
        'IT00000000000000000000000000000087': 'Dark Brown Hair and Beard Dye',
        'IT00000000000000000000000000000088': 'Natural Black Hair and Beard Dye',
        'IT00000000000000000000000000000089': 'Jet Black Hair and Beard Dye',
        'IT00000000000000000000002001000001': 'Barrier Cream Packette',
        'IT00000000000000000000002001000002': 'Safety Instructions',
        'IT00000000000000000000002001000003': 'Face Wipe',
        'IT00000000000000000000002001000004': 'Customized Instructions',
        'IT00000000000000000000000000000090': 'Sensitive Face & Beard Scrub',
        'IT00000000000000000000000000000091': 'Energizing Face & Beard Scrub',
        'IT00000000000000000000000000000092': 'Cleverman Face Scrubber',
        'IT00000000000000000000000000000093': 'Cleverman Tweezer',
        'IT00000000000000000000000000000094': 'Cleverman Grooming Scissors',
        'IT00000000000000000000000000000095': 'Cleverman Fingernail and Toenail Clippers',
        'IT00000000000000000000000000000096': 'Cleverman Shampoo 20ml Thin',
        'IT00000000000000000000000000000097': 'Cleverman Conditioner 20ml - Damaged',
        'IT00000000000000000000000000000098': 'Dark Blond Hair and Beard Dye',
        'IT00000000000000000000000000000099': 'Cleverman Face Wipe - 10 Pack',
        'IT00000000000000000000000000000100': 'Cleverman Normal Shampoo - 5 Pack',
        'IT00000000000000000000000000000101': 'Cleverman Damaged Conditioner - 5 Pack',
        'IT00000000000000000000000000000102': 'Black Hair and Beard Dye',
        'IT00000000000000000000000000000103': 'Beard Scrub 7ml Sensitive',
        'IT00000000000000000000000000000104': 'Beard Scrub 7ml Energizing',
        'IT00000000000000000000000000000105': 'Jet Black Hair and Beard Dye FBM',
        'IT00000000000000000000000000000106': 'Dark Brown Hair and Beard Dye FBM',
        'IT00000000000000000000000000000107': 'Jaxon Lane Boom Cica Wow Cleanser',
        'IT00000000000000000000001004170008': 'Full Coverage Immediate 30ml Kit',
        'IT00000000000000000000001004170009': 'Full Coverage Gradual 30ml Kit',
        'IT00000000000000000000001004170010': 'Salt & Pepper Targeted 30ml Kit',
        'IT00000000000000000000000000000108': 'Triptych Brochure for OTO Beard Kits',
        'IT00000000000000000000000000000109': 'Card - All Subscription Orders',
        'IT00000000000000000000000000000110': 'Moisturizing Beard Scrub Single Use Sachet',
        'IT00000000000000000000000000000111': 'Energizing Beard Scrub Single Use Sachet'
    }

    # Diccionario de clasificación basado en IDs
    product_classification = {
        # Kit components
        'IT00000000000000000000000000000001': "Kit component",
        'IT00000000000000000000000000000002': "Kit component",
        'IT00000000000000000000000000000005': "Kit component",
        'IT00000000000000000000000000000006': "Kit component",
        'IT00000000000000000000000000000007': "Kit component",
        'IT00000000000000000000000000000008': "Kit component",
        'IT00000000000000000000000000000009': "Kit component",
        'IT00000000000000000000000000000010': "Kit component",
        'IT00000000000000000000000000000011': "Kit component",
        'IT00000000000000000000000000000012': "Kit component",
        'IT00000000000000000000000000000013': "Kit component",
        'IT00000000000000000000000000000014': "Kit component",
        'IT00000000000000000000000000000015': "Kit component",
        'IT00000000000000000000000000000016': "Kit component",
        'IT00000000000000000000000000000017': "Kit component",
        'IT00000000000000000000000000000018': "Kit component",
        'IT00000000000000000000000000000019': "Kit component",
        'IT00000000000000000000000000000020': "Kit component",
        'IT00000000000000000000000000000021': "Kit component",
        'IT00000000000000000000000000000022': "Kit component",
        'IT00000000000000000000000000000023': "Kit component",
        'IT00000000000000000000000000000024': "Kit component",
        'IT00000000000000000000000000000025': "Kit component",
        'IT00000000000000000000000000000026': "Kit component",
        'IT00000000000000000000000000000028': "Kit component",
        'IT00000000000000000000000000000030': "Kit component",
        'IT00000000000000000000000000000031': "Kit component",
        'IT00000000000000000000000000000032': "Kit component",
        'IT00000000000000000000000000000033': "Kit component",
        'IT00000000000000000000000000000034': "Kit component",
        'IT00000000000000000000000000000035': "Kit component",
        'IT00000000000000000000000000000036': "Kit component",
        'IT00000000000000000000000000000037': "Kit component",
        'IT00000000000000000000000000000038': "Kit component",
        'IT00000000000000000000000000000039': "Kit component",
        'IT00000000000000000000000000000040': "Kit component",
        'IT00000000000000000000000000000041': "Kit component",
        'IT00000000000000000000000000000042': "Kit component",
        'IT00000000000000000000000000000043': "Kit component",
        'IT00000000000000000000000000000044': "Kit component",
        'IT00000000000000000000000000000045': "Kit component",
        'IT00000000000000000000000000000046': "Kit component",
        'IT00000000000000000000000000000047': "Kit component",
        'IT00000000000000000000000000000048': "Kit component",
        'IT00000000000000000000000000000049': "Kit component",
        'IT00000000000000000000000000000050': "Kit component",
        'IT00000000000000000000000000000051': "Kit component",
        'IT00000000000000000000000000000052': "Kit component",
        'IT00000000000000000000000000000053': "Kit component",
        'IT00000000000000000000000000000054': "Kit component",
        'IT00000000000000000000000000000055': "Kit component",
        'IT00000000000000000000000000000056': "Kit component",
        'IT00000000000000000000000000000057': "Kit component",
        'IT00000000000000000000000000000058': "Kit component",
        'IT00000000000000000000000000000059': "Kit component",
        'IT00000000000000000000000000000060': "Kit component",
        'IT00000000000000000000000000000061': "Kit component",
        'IT00000000000000000000000000000062': "Kit component",
        'IT00000000000000000000000000000063': "Kit component",
        'IT00000000000000000000000000000064': "Kit component",
        'IT00000000000000000000000000000065': "Kit component",
        'IT00000000000000000000000000000066': "Kit component",
        'IT00000000000000000000000000000067': "Kit component",
        'IT00000000000000000000000000000068': "Kit component",
        'IT00000000000000000000000000000069': "Kit component",
        'IT00000000000000000000000000000070': "Kit component",
        'IT00000000000000000000000000000071': "Kit component",
        'IT00000000000000000000000000000072': "Kit component",
        'IT00000000000000000000000000000073': "Kit component",
        'IT00000000000000000000000000000074': "Kit component",
        'IT00000000000000000000000000000075': "Kit component",
        'IT00000000000000000000000000000076': "Kit component",
        'IT00000000000000000000000000000004': "Kit component",
        'IT00000000000000000000000000000003': "Kit component",
        'IT00000000000000000000000000000029': "Kit component",
        'IT00000000000000000000000000000027': "Kit component",
        'IT00000000000000000000000000000103': "Kit component",  # Beard Scrub 7ml Sensitive
        'IT00000000000000000000000000000104': "Kit component",  # Beard Scrub 7ml Energizing
        'IT00000000000000000000000000000097': "Kit component",
        'IT00000000000000000000000000000096': "Kit component",
        'IT00000000000000000000000000000081': "Kit component",
        
        # All in one
        'IT00000000000000000000000000000082': "All in one",  # Lightest Blond
        'IT00000000000000000000000000000083': "All in one",  # Medium Blond
        'IT00000000000000000000000000000084': "All in one",  # Auburn
        'IT00000000000000000000000000000085': "All in one",  # Light Brown
        'IT00000000000000000000000000000086': "All in one",  # Medium Brown
        'IT00000000000000000000000000000087': "All in one",  # Dark Brown
        'IT00000000000000000000000000000088': "All in one",  # Natural Black
        'IT00000000000000000000000000000089': "All in one",  # Jet Black
        'IT00000000000000000000000000000098': "All in one",  # Dark Blond
        'IT00000000000000000000000000000102': "All in one",  # Black
        'IT00000000000000000000000000000105': "All in one",  # Jet Black FBM
        'IT00000000000000000000000000000106': "All in one",  # Dark Brown FBM
        
        # Custom kits
        'IT00000000000000000000001004170001': "Custom kit",
        'IT00000000000000000000001004170002': "Custom kit",
        'IT00000000000000000000001004170003': "Custom kit",
        'IT00000000000000000000001004170004': "Custom kit",
        'IT00000000000000000000001004170005': "Custom kit",
        'IT00000000000000000000001004170006': "Custom kit",
        'IT00000000000000000000001004170007': "Custom kit",
        'IT00000000000000000000001004170008': "Custom kit",
        'IT00000000000000000000001004170009': "Custom kit",
        'IT00000000000000000000001004170010': "Custom kit",
        
        # NPD
        'IT00000000000000000000000000000090': "NPD",
        'IT00000000000000000000000000000091': "NPD",
        'IT00000000000000000000000000000092': "NPD",
        'IT00000000000000000000000000000093': "NPD",
        'IT00000000000000000000000000000094': "NPD",
        'IT00000000000000000000000000000095': "NPD",
        'IT00000000000000000000000000000099': "NPD",
        'IT00000000000000000000000000000100': "NPD",
        'IT00000000000000000000000000000101': "NPD",
        
        # Otros
        'IT00000000000000000000000000000079': "Otros",
        'IT00000000000000000000000000000080': "Otros",
        'IT00000000000000000000002001000001': "Otros",
        'IT00000000000000000000002001000002': "Otros",
        'IT00000000000000000000002001000003': "Otros",
        'IT00000000000000000000002001000004': "Otros",
        'IT00000000000000000000000000000107': "Otros",
        'IT00000000000000000000000000000108': "Otros",
        'IT00000000000000000000000000000109': "Otros",
        'IT00000000000000000000000000000110': "Otros",
        'IT00000000000000000000000000000111': "Otros"
    }

    return id_to_name, product_classification

def analyze_product_rebuys():
    """Genera el reporte de recompras por tipo de producto"""
    start_date, end_date = open_rebuy_date_selector()
    if not start_date or not end_date:
        return

    # Consulta SQL modificada para incluir solo usuarios cuya primera compra está en el rango
    rebuy_query = f"""
    WITH first_time_buyers AS (
        -- Usuarios que hicieron su primera compra en el rango de fechas
        SELECT customer_id
        FROM bi.fact_orders
        WHERE status != 'CANCELLED'
        AND order_plan = 'OTO'
        AND is_first_order = 1
        AND created_at BETWEEN '{start_date}' AND '{end_date}'
    ),
    filtered_orders AS (
        -- Todas las órdenes de esos usuarios (incluyendo sus recompras)
        SELECT fo.id, fo.customer_id
        FROM bi.fact_orders fo
        JOIN first_time_buyers ftb ON fo.customer_id = ftb.customer_id
        WHERE fo.status != 'CANCELLED'
        AND fo.order_plan = 'OTO'
        AND fo.recurrent = 0
    ),
    customer_product_orders AS (
        -- Conteo de órdenes por cliente y producto
        SELECT 
            fo.customer_id,
            soi.itemId,
            COUNT(DISTINCT fo.id) as order_count
        FROM filtered_orders fo
        JOIN bi.fact_sales_order_items soi ON fo.id = soi.salesOrderId
        GROUP BY fo.customer_id, soi.itemId
    ),
    product_order_counts AS (
        -- Total de órdenes por producto (solo de usuarios que cumplen el criterio)
        SELECT 
            soi.itemId,
            COUNT(DISTINCT fo.id) as total_orders
        FROM filtered_orders fo
        JOIN bi.fact_sales_order_items soi ON fo.id = soi.salesOrderId
        GROUP BY soi.itemId
    ),
    product_customer_counts AS (
        -- Total de clientes únicos por producto (solo los que cumplen)
        SELECT 
            soi.itemId,
            COUNT(DISTINCT fo.customer_id) as total_customers
        FROM filtered_orders fo
        JOIN bi.fact_sales_order_items soi ON fo.id = soi.salesOrderId
        GROUP BY soi.itemId
    ),
    product_customer_single_purchase AS (
        -- Clientes que solo compraron una vez (dentro del criterio)
        SELECT 
            soi.itemId,
            COUNT(DISTINCT fo.customer_id) as single_purchase_customers
        FROM filtered_orders fo
        JOIN bi.fact_sales_order_items soi ON fo.id = soi.salesOrderId
        WHERE fo.customer_id IN (
            SELECT customer_id
            FROM filtered_orders
            GROUP BY customer_id
            HAVING COUNT(DISTINCT id) = 1
        )
        GROUP BY soi.itemId
    )
    SELECT 
        cpo.itemId as item_id,
        i.name as Producto,
        cpo.customer_id,
        CASE WHEN cpo.order_count > 1 THEN cpo.order_count - 1 ELSE 0 END as rebuy_count,
        poc.total_orders,
        pcc.total_customers,
        COALESCE(psp.single_purchase_customers, 0) as single_purchase_customers
    FROM customer_product_orders cpo
    JOIN bi.dim_items i ON cpo.itemId = i.id
    JOIN product_order_counts poc ON cpo.itemId = poc.itemId
    JOIN product_customer_counts pcc ON cpo.itemId = pcc.itemId
    LEFT JOIN product_customer_single_purchase psp ON cpo.itemId = psp.itemId
    """

    # Obtener datos y aplicar clasificación manual
    rebuy_data = execute_query(rebuy_query)
    id_to_name, product_classification = get_product_classification()
    
    # Asignar nombres basados en IDs primero
    rebuy_data['Producto'] = rebuy_data['item_id'].map(id_to_name)
    
    # Asignar tipos basados en IDs
    rebuy_data['Tipo'] = rebuy_data['item_id'].map(product_classification)
    rebuy_data = rebuy_data.dropna(subset=['Tipo'])

    # Crear Excel con una hoja por tipo
    output_file = f"Reporte_Recompras_{start_date}_a_{end_date}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja vacía inicial

    for product_type in rebuy_data['Tipo'].unique():
        # Filtrar datos por tipo
        type_data = rebuy_data[rebuy_data['Tipo'] == product_type]
        
        # Calcular máx. recompras para este tipo
        max_rebuy = type_data['rebuy_count'].max()
        
        # Crear tabla dinámica
        pivot = type_data.groupby(['Producto', 'rebuy_count']).size().unstack(fill_value=0)
        pivot = pivot.reindex(columns=range(0, max_rebuy + 1), fill_value=0)  # Ahora incluye columna 0
        
        # Añadir columna 0 (single purchase) si no existe
        if 0 not in pivot.columns:
            pivot[0] = 0
        
        # Ordenar columnas numéricamente (0, 1, 2, ...)
        pivot = pivot.reindex(sorted(pivot.columns), axis=1)
        pivot.columns = [str(col) for col in pivot.columns]  # Columnas: 0, 1, 2...
        
        # Añadir métricas adicionales
        pivot['Total usuarios que recompraron'] = pivot[[col for col in pivot.columns if col != '0']].sum(axis=1)
        pivot['Total compras'] = type_data.groupby('Producto')['total_orders'].first()
        pivot['Total usuarios'] = type_data.groupby('Producto')['total_customers'].first()
        pivot['% Recompra'] = (pivot['Total usuarios que recompraron'] / pivot['Total usuarios']).round(4)
        
        # Reordenar columnas como solicitado
        pivot = pivot[['Total usuarios que recompraron', 'Total compras', 'Total usuarios', '% Recompra'] + 
                     [col for col in pivot.columns if col not in ['Total usuarios que recompraron', 'Total compras', 'Total usuarios', '% Recompra']]]
        
        # Crear hoja en Excel
        ws = wb.create_sheet(title=product_type[:31])  # Máx. 31 caracteres
        
        # Escribir encabezados
        headers = ['Producto'] + list(pivot.columns)
        ws.append(headers)
        
        # Escribir datos
        for product, row in pivot.iterrows():
            ws.append([product] + list(row))
        
        # Ajustar anchos de columnas
        for col in ws.columns:
            max_len = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = (max_len + 2) * 1.2
        
        # Formato porcentaje y estilo
        percent_col = headers.index('% Recompra') + 1  # +1 porque Excel es base 1
        for row in ws.iter_rows(min_row=2, min_col=percent_col, max_col=percent_col):
            for cell in row:
                cell.number_format = '0.00%'
        
        # Congelar encabezados
        ws.freeze_panes = 'A2'

    # Guardar y notificar
    wb.save(output_file)
    messagebox.showinfo("Listo", f"Reporte guardado como:\n{output_file}")

if __name__ == "__main__":
    analyze_product_rebuys()