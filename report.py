import os
from openpyxl import load_workbook
import tkinter as tk
from tkinter import BooleanVar, Checkbutton, Button

# Cargar el archivo Excel existente


hoja_report = "Report"  
hoja_files = "Files" 


def anotar_datos_excel(datos, columna_inicio, fila_inicio, urls=False, month = '', primer_uso=False):
    
    archivo_excel = 'Monthly Report.xlsx'
    nuevo_archivo = f'Monthly Report {month}.xlsx'

    if primer_uso == False:
        archivo_excel = nuevo_archivo

    try:
        # Intentar cargar el archivo Excel existente
        wb = load_workbook(archivo_excel)
    except FileNotFoundError:
        print(f"El archivo '{archivo_excel}' no existe.")
        return

    # Seleccionar o crear la hoja donde escribirás los datos
    if(urls):
        ws = wb[hoja_files]
    else:
        ws = wb[hoja_report]

    # Escribir los datos en las celdas
    for i, valor in enumerate(datos, start=fila_inicio):
        ws.cell(row=i, column=columna_inicio, value=valor)

    # Guardar los cambios en el archivo (nuevo o existente)
    wb.save(nuevo_archivo)

def seleccionar_tipo_de_reporte():
    """
    Muestra una ventana para que el usuario seleccione el tipo de reporte a generar.

    :return: tuple (funnels_report, database_report) - Estado de los checkboxes seleccionados por el usuario.
    """
    # Variables para almacenar el estado de los checkboxes
    root = tk.Tk()
    root.title("Seleccionar Tipo de Reporte")

    funnels_var = BooleanVar(value=True)
    database_var = BooleanVar(value=True)
    block_payments_var = BooleanVar(value=True)

    # Etiqueta principal
    label = tk.Label(root, text="Seleccione el tipo de reporte")
    label.pack(pady=10)

    # Checkbox para Funnels
    funnel_checkbox = Checkbutton(root, text="Funnels", variable=funnels_var)
    funnel_checkbox.pack(anchor=tk.W, padx=20)

    # Checkbox para Base de Datos
    database_checkbox = Checkbutton(root, text="Base de Datos", variable=database_var)
    database_checkbox.pack(anchor=tk.W, padx=20)

     # Checkbox para Base de Datos
    block_payments_checkbox = Checkbutton(root, text="Pagos bloqueados", variable=block_payments_var)
    block_payments_checkbox.pack(anchor=tk.W, padx=20)

    # Función para cerrar la ventana y devolver los valores seleccionados
    def continuar():
        root.quit()
        root.destroy()

    # Botón para continuar
    continuar_button = Button(root, text="Continuar", command=continuar)
    continuar_button.pack(pady=20)

    root.mainloop()

    return funnels_var.get(), database_var.get(), block_payments_var.get()

def seleccionar_donde_almacenar():
    """
    Muestra una ventana para que el usuario seleccione donde almacenar el reporte.
    Los checkboxes son mutuamente excluyentes (solo se puede seleccionar uno).

    :return: tuple (dropbox_var, drive_var) - Estado de los checkboxes seleccionados por el usuario.
    """
    # Variables para almacenar el estado de los checkboxes
    root = tk.Tk()
    root.title("Guardar reporte en la nube")

    dropbox_var = BooleanVar(value=False)
    drive_var = BooleanVar(value=False)

    # Función para hacer los checkboxes mutuamente excluyentes
    def toggle_checkboxes(selected_var, other_var):
        if selected_var.get():  # Si el checkbox actual fue seleccionado
            other_var.set(False)  # Deseleccionar el otro

    # Etiqueta principal
    label = tk.Label(root, text="Seleccione el tipo de reporte")
    label.pack(pady=10)

    # Checkbox para Dropbox
    dropbox_checkbox = Checkbutton(
        root, 
        text="Dropbox", 
        variable=dropbox_var,
        command=lambda: toggle_checkboxes(dropbox_var, drive_var)
    )
    dropbox_checkbox.pack(anchor=tk.W, padx=20)

    # Checkbox para Google Drive
    drive_checkbox = Checkbutton(
        root, 
        text="Google drive", 
        variable=drive_var,
        command=lambda: toggle_checkboxes(drive_var, dropbox_var)
    )
    drive_checkbox.pack(anchor=tk.W, padx=20)

    # Función para cerrar la ventana y devolver los valores seleccionados
    def continuar():
        root.quit()
        root.destroy()

    # Botón para continuar
    continuar_button = Button(root, text="Continuar", command=continuar)
    continuar_button.pack(pady=20)

    root.mainloop()

    return dropbox_var.get(), drive_var.get()