import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
from modules.database_queries import execute_query
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO

# === CONSULTAS ===
query_descuentos = """
SELECT 
    fo.id,
    fo.created_at
FROM prod_sales_and_subscriptions.discounts ds
JOIN bi.fact_orders fo ON ds.entityId = fo.id
WHERE ds.name LIKE '3x2%'
  AND ds.createdAt > '2024-10-01'
  AND ds.createdAt < '2025-10-01'
  AND fo.status != 'CANCELLED'
  AND fo.order_plan != 'SUBSCRIPTION';
"""

query_ordenes = """
SELECT 
    id,
    created_at
FROM bi.fact_orders
WHERE recurrent = 0
  AND created_at > '2024-10-01'
  AND created_at < '2025-10-01'
  AND status != 'CANCELLED'
  AND order_plan != 'SUBSCRIPTION';
"""

# === EJECUCIÓN DE CONSULTAS ===
df_desc = execute_query(query_descuentos)
df_ord  = execute_query(query_ordenes)

# Fechas a datetime
df_desc['created_at'] = pd.to_datetime(df_desc['created_at'])
df_ord['created_at']  = pd.to_datetime(df_ord['created_at'])

# === AGRUPACIÓN POR DÍA (TODO EL PERÍODO) ===
desc_por_dia_all = (
    df_desc.assign(created_day=df_desc['created_at'].dt.floor('D'))
           .groupby('created_day').size()
           .reset_index(name='descuentos')
           .rename(columns={'created_day':'created_at'})
)

ord_por_dia_all = (
    df_ord.assign(created_day=df_ord['created_at'].dt.floor('D'))
          .groupby('created_day').size()
          .reset_index(name='ordenes')
          .rename(columns={'created_day':'created_at'})
)

# === FILTRO PARA GRÁFICAS POR DÍA: ÚLTIMOS 2 MESES ===
fecha_limite_ts = pd.Timestamp.now().normalize() - pd.DateOffset(months=2)
desc_por_dia_2m = desc_por_dia_all[desc_por_dia_all['created_at'] >= fecha_limite_ts].copy()
ord_por_dia_2m  = ord_por_dia_all[ord_por_dia_all['created_at']  >= fecha_limite_ts].copy()

if desc_por_dia_2m.empty:
    desc_por_dia_2m = desc_por_dia_all.copy()
if ord_por_dia_2m.empty:
    ord_por_dia_2m = ord_por_dia_all.copy()

# === AGRUPACIÓN POR MES (TODO EL PERÍODO) ===
desc_por_mes = (
    df_desc.groupby(df_desc['created_at'].dt.to_period('M')).size()
           .reset_index(name='descuentos')
    .assign(mes=lambda d: d['created_at'].astype(str))
    .drop(columns=['created_at'])
)

ord_por_mes = (
    df_ord.groupby(df_ord['created_at'].dt.to_period('M')).size()
          .reset_index(name='ordenes')
    .assign(mes=lambda d: d['created_at'].astype(str))
    .drop(columns=['created_at'])
)

# === PROMEDIOS ===
prom_dia_desc = desc_por_dia_all['descuentos'].mean() if not desc_por_dia_all.empty else 0
prom_mes_desc = desc_por_mes['descuentos'].mean() if not desc_por_mes.empty else 0
prom_dia_ord  = ord_por_dia_all['ordenes'].mean() if not ord_por_dia_all.empty else 0
prom_mes_ord  = ord_por_mes['ordenes'].mean() if not ord_por_mes.empty else 0

# === % ÓRDENES CON DESCUENTOS POR DÍA ===
comparacion_all = pd.merge(
    ord_por_dia_all, desc_por_dia_all,
    on='created_at', how='left'
).fillna({'descuentos': 0})

comparacion_all['pct_con_descuento'] = (
    (comparacion_all['descuentos'] / comparacion_all['ordenes']) * 100
).replace([np.inf, -np.inf], 0).fillna(0)

# === FUNCIÓN: gráfico con línea de regresión ===
def crear_grafico(df, x_col, y_col, titulo, ancho, color_bar='steelblue'):
    fig, ax = plt.subplots(figsize=(ancho, 4))
    # Ordenar por X si aplica
    if np.issubdtype(df[x_col].dtype, np.datetime64):
        df = df.sort_values(x_col)
        x_labels = df[x_col].dt.strftime('%Y-%m-%d')
    else:
        x_labels = df[x_col].astype(str)

    y = df[y_col].values
    x = np.arange(len(df))

    ax.bar(x_labels, y, color=color_bar, alpha=0.7)
    ax.set_title(titulo)
    ax.set_xlabel('')
    ax.set_ylabel('Cantidad')
    plt.xticks(rotation=45)

    if len(x) > 1 and np.any(y):
        m, b = np.polyfit(x, y, 1)
        ax.plot(x_labels, m * x + b, color='red', linewidth=2, label='Tendencia')
        ax.legend()

    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format='png')
    plt.close(fig)
    buf.seek(0)
    return buf

# === GRÁFICOS ===
img_dia_desc = crear_grafico(desc_por_dia_2m, 'created_at', 'descuentos',
                             'Descuentos por Día (últimos 2 meses)', 30)
img_dia_ord  = crear_grafico(ord_por_dia_2m,  'created_at', 'ordenes',
                             'Órdenes por Día (últimos 2 meses)', 30)
img_mes_desc = crear_grafico(desc_por_mes, 'mes', 'descuentos',
                             'Descuentos por Mes (todo el período)', 8)
img_mes_ord  = crear_grafico(ord_por_mes,  'mes', 'ordenes',
                             'Órdenes por Mes (todo el período)', 8)

# === EXPORTAR A EXCEL ===
wb = Workbook()

# --- Hoja Descuentos ---
ws_desc = wb.active
ws_desc.title = "Descuentos"
for r in dataframe_to_rows(desc_por_dia_all, index=False, header=True):
    ws_desc.append(r)
ws_desc.append([])
ws_desc.append(["Promedio por día (todo el período)", prom_dia_desc])
ws_desc.append(["Promedio por mes (todo el período)", prom_mes_desc])
ws_desc.add_image(XLImage(img_dia_desc), "E2")
ws_desc.add_image(XLImage(img_mes_desc), "E22")

# --- Hoja Órdenes ---
ws_ord = wb.create_sheet("Órdenes")
for r in dataframe_to_rows(ord_por_dia_all, index=False, header=True):
    ws_ord.append(r)
ws_ord.append([])
ws_ord.append(["Promedio por día (todo el período)", prom_dia_ord])
ws_ord.append(["Promedio por mes (todo el período)", prom_mes_ord])
ws_ord.add_image(XLImage(img_dia_ord), "E2")
ws_ord.add_image(XLImage(img_mes_ord), "E22")

# --- Hoja Comparación diaria ---
ws_cmp = wb.create_sheet("Comparación (diario)")
for r in dataframe_to_rows(comparacion_all, index=False, header=True):
    ws_cmp.append(r)

# === RESUMEN MENSUAL ===
comparacion_all['mes_period'] = comparacion_all['created_at'].dt.to_period('M')
agg = comparacion_all.groupby('mes_period').agg(
    ordenes_sum=('ordenes', 'sum'),
    descuentos_sum=('descuentos', 'sum'),
    dias=('created_at', 'nunique')
).reset_index()

agg['prom_ordenes_dia'] = agg['ordenes_sum'] / agg['dias']
agg['prom_desc_dia']    = agg['descuentos_sum'] / agg['dias']
agg['pct_con_descuento_mes'] = np.where(
    agg['ordenes_sum'] > 0,
    (agg['descuentos_sum'] / agg['ordenes_sum']) * 100,
    0
)

# Totales globales
total_ordenes   = agg['ordenes_sum'].sum()
total_desc      = agg['descuentos_sum'].sum()
total_dias_dist = comparacion_all['created_at'].dt.date.nunique()

fila_total = pd.DataFrame({
    'mes_period': ['TOTAL'],
    'ordenes_sum': [total_ordenes],
    'descuentos_sum': [total_desc],
    'dias': [total_dias_dist],
    'prom_ordenes_dia': [ (total_ordenes / total_dias_dist) if total_dias_dist else 0 ],
    'prom_desc_dia': [ (total_desc / total_dias_dist) if total_dias_dist else 0 ],
    'pct_con_descuento_mes': [ (total_desc / total_ordenes * 100) if total_ordenes else 0 ]
})

resumen_final = pd.concat([agg, fila_total], ignore_index=True)
resumen_final['mes'] = resumen_final['mes_period'].astype(str)
resumen_final = resumen_final.drop(columns=['mes_period'])

# --- Hoja resumen ---
ws_res = wb.create_sheet("Resumen mensual")
ws_res.append(["Resumen mensual (todo el período)"])
ws_res.append([])
for r in dataframe_to_rows(resumen_final, index=False, header=True):
    ws_res.append(r)

# === Gráfico de % mensual con regresión ===
img_pct_mes = crear_grafico(resumen_final[resumen_final['mes'] != 'TOTAL'],
                            'mes', 'pct_con_descuento_mes',
                            '% de Órdenes con Descuento por Mes', 10, color_bar='orange')

ws_res.add_image(XLImage(img_pct_mes), "E5")

# --- Guardar archivo ---
nombre_archivo = f"analisis_descuentos_ordenes_{datetime.now().strftime('%Y%m%d')}.xlsx"
wb.save(nombre_archivo)

print(f"Archivo Excel generado: {nombre_archivo}")
