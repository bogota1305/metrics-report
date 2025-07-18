import pandas as pd
import tkinter as tk
from tkinter import Label, Button, messagebox, Listbox, END, MULTIPLE, Frame
from tkcalendar import Calendar

from modules.database_queries import execute_query

# Diccionario de productos (el que proporcionaste)
product_dict = {
    'IT00000000000000000000000000000001': 'Big Shipper: Small Gloves',
        'IT00000000000000000000000000000002': '30ml Colorant - Medium Brown',
        'IT00000000000000000000000000000005': 'Small Shipper: X-Large Gloves',
        'IT00000000000000000000000000000006': '30ml Colorant - Jet-Black',
        'IT00000000000000000000000000000007': '30ml Colorant - Black',
        'IT00000000000000000000000000000008': '30ml Colorant - Soft-Black',
        'IT00000000000000000000000000000009': '30ml Colorant - Dark Brown',
        'IT00000000000000000000000000000010': '30ml Colorant - Medium-Dark Brown',
        'IT00000000000000000000000000000011': '30ml Colorant - Medium-Light Brown',
        'IT00000000000000000000000000000012': '30ml Colorant - Light Brown',
        'IT00000000000000000000000000000013': '30ml Colorant - Dark Blond',
        'IT00000000000000000000000000000014': '30ml Colorant - Warm- Dark Blond',
        'IT00000000000000000000000000000015': '30ml Colorant - Cool- Medium Blond',
        'IT00000000000000000000000000000016': '30ml Colorant - Warm- Medium Blond',
        'IT00000000000000000000000000000017': '30ml Colorant - Cool- Light Blond',
        'IT00000000000000000000000000000018': '30ml Colorant - Warm- Light Blond',
        'IT00000000000000000000000000000019': '30ml Colorant - Lightest Blond',
        'IT00000000000000000000000000000020': '30ml Colorant - Dark Auburn',
        'IT00000000000000000000000000000021': '30ml Colorant - Light Auburn',
        'IT00000000000000000000000000000022': '30ml Developer - 20 Vol',
        'IT00000000000000000000000000000023': '30ml Developer - 10 Vol',
        'IT00000000000000000000000000000024': 'Tray with Spatula',
        'IT00000000000000000000000000000025': '7 Min: Short Beard Brush',
        'IT00000000000000000000000000000026': '10 Min: Short Beard Brush',
        'IT00000000000000000000000000000028': '10 Min: Medium Beard Brush',
        'IT00000000000000000000000000000030': '10 Min: Long Beard Brush',
        'IT00000000000000000000000000000031': 'Small Shipper: Normal Hair Treatment',
        'IT00000000000000000000000000000032': 'Small Shipper: Thin Hair Treatment',
        'IT00000000000000000000000000000033': 'Small Shipper: Ultra Dry Hair Treatment',
        'IT00000000000000000000000000000034': 'Small Shipper: Damaged Hair Treatment',
        'IT00000000000000000000000000000035': 'Small Shipper: Coily Hair Treatment',
        'IT00000000000000000000000000000036': 'Small Shipper: Moisturizing',
        'IT00000000000000000000000000000037': 'Small Shipper: Sensitive',
        'IT00000000000000000000000000000038': 'Small Shipper: Energizing',
        'IT00000000000000000000000000000039': 'Small Shipper',
        'IT00000000000000000000000000000040': 'Big Shipper',
        'IT00000000000000000000000000000041': 'Small Shipper: Small Gloves',
        'IT00000000000000000000000000000042': 'Big Shipper: Medium Gloves',
        'IT00000000000000000000000000000043': 'Big Shipper: Large Gloves',
        'IT00000000000000000000000000000044': 'Big Shipper: X-Large Gloves',
        'IT00000000000000000000000000000045': '45ml Colorant - Jet-Black',
        'IT00000000000000000000000000000046': '45ml Colorant - Black',
        'IT00000000000000000000000000000047': '45ml Colorant - Soft-Black',
        'IT00000000000000000000000000000048': '45ml Colorant - Dark Brown',
        'IT00000000000000000000000000000049': '45ml Colorant - Medium-Dark Brown',
        'IT00000000000000000000000000000050': '45ml Colorant - Medium Brown',
        'IT00000000000000000000000000000051': '45ml Colorant - Medium-Light Brown',
        'IT00000000000000000000000000000052': '45ml Colorant: Light Brown',
        'IT00000000000000000000000000000053': '45ml Colorant - Dark Blond',
        'IT00000000000000000000000000000054': '45ml Colorant - Warm-Dark Blond',
        'IT00000000000000000000000000000055': '45ml Colorant - Cool-Medium Blond',
        'IT00000000000000000000000000000056': '45ml Colorant - Warm-Medium Blond',
        'IT00000000000000000000000000000057': '45ml Colorant - Cool - Light Blond',
        'IT00000000000000000000000000000058': '45ml Colorant - Warm - Light Blond',
        'IT00000000000000000000000000000059': '45ml Colorant - Lightest Blond',
        'IT00000000000000000000000000000060': '45ml Colorant - Dark Auburn',
        'IT00000000000000000000000000000061': '45ml Colorant - Light Auburn',
        'IT00000000000000000000000000000062': '45ml Developer - 30 Vol',
        'IT00000000000000000000000000000063': '45ml Developer - 20 Vol',
        'IT00000000000000000000000000000064': '5 Min: Big Shipper - Brush and Tray',
        'IT00000000000000000000000000000065': '7 Min: Big Shipper - Brush and Tray',
        'IT00000000000000000000000000000066': '10 Min: Big Shipper - Brush and Tray',
        'IT00000000000000000000000000000067': '5 Min: Big Shipper - Brush+Tray+Spatula',
        'IT00000000000000000000000000000068': '7 Min: Big Shipper - Brush+Tray+Spatula',
        'IT00000000000000000000000000000069': '10 Min: Big Shipper - Brush+Tray+Spatula',
        'IT00000000000000000000000000000070': 'Big Shipper: Normal Hair Treatment',
        'IT00000000000000000000000000000071': 'Big Shipper: Thin Hair Treatment',
        'IT00000000000000000000000000000072': 'Big Shipper: Ultra Dry Hair Treatment',
        'IT00000000000000000000000000000073': 'Big Shipper: Damaged Hair Treatment',
        'IT00000000000000000000000000000074': 'Big Shipper: 4C / Afro-Textured Hair',
        'IT00000000000000000000000000000075': 'Box 2 MIX Empty - Big Shipper',
        'IT00000000000000000000000000000076': 'Box 2 MIX Pre-Packed - Small Shipper',
        'IT00000000000000000000000000000004': 'Small Shipper: Large Gloves',
        'IT00000000000000000000000000000003': 'Small Shipper: Medium Gloves',
        'IT00000000000000000000000000000029': '7 Min: Long Beard Brush',
        'IT00000000000000000000000000000027': '7 Min: Medium Beard Brush',
        'IT00000000000000000000000000000103': 'Beard Scrub 7ml Sensitive',
        'IT00000000000000000000000000000104': 'Beard Scrub 7ml Energizing',
}

def open_rebuy_date_selector():
    """Interfaz para seleccionar fechas de análisis"""
    start_date = None
    end_date = None
    
    def get_dates():
        nonlocal start_date, end_date
        start_date = cal_start.get_date()
        end_date = cal_end.get_date()
        
        if not start_date or not end_date:
            messagebox.showerror("Error", "Selecciona ambas fechas.")
            return
            
        # Convertir a formato YYYY-MM-DD
        start_date = pd.to_datetime(start_date).strftime('%Y-%m-%d')
        end_date = pd.to_datetime(end_date).strftime('%Y-%m-%d')
        
        root.quit()
        root.destroy()

    # Configuración de la ventana
    root = tk.Tk()
    root.title("Selección de Rango de Fechas")
    
    Label(root, text="Fecha de inicio:").pack(pady=5)
    cal_start = Calendar(root, selectmode='day', date_pattern='yyyy-mm-dd')
    cal_start.pack(pady=5)

    Label(root, text="Fecha de fin:").pack(pady=5)
    cal_end = Calendar(root, selectmode='day', date_pattern='yyyy-mm-dd')
    cal_end.pack(pady=5)

    Button(root, text="Generar Reporte", command=get_dates).pack(pady=20)
    
    root.mainloop()
    return start_date, end_date

def select_report_type():
    """Interfaz para seleccionar tipo de reporte"""
    report_types = []
    
    def get_selection():
        nonlocal report_types
        report_types = []
        if var_total.get():
            report_types.append('total')
        if var_combinations.get():
            report_types.append('combinations')
        
        if not report_types:
            messagebox.showerror("Error", "Selecciona al menos un tipo de reporte")
            return
            
        root.quit()
        root.destroy()

    root = tk.Tk()
    root.title("Selección de Tipo de Reporte")
    
    var_total = tk.BooleanVar()
    var_combinations = tk.BooleanVar()
    
    tk.Label(root, text="Selecciona el tipo de reporte:").pack(pady=10)
    tk.Checkbutton(root, text="Reporte Total (por producto individual)", variable=var_total).pack(anchor='w')
    tk.Checkbutton(root, text="Reporte de Combinaciones", variable=var_combinations).pack(anchor='w')
    
    Button(root, text="Continuar", command=get_selection).pack(pady=20)
    
    root.mainloop()
    return report_types

def select_product_combinations():
    """Interfaz para seleccionar combinaciones de productos"""
    display_names = sorted(set(product_dict.values()))
    selected_combinations = []
    
    def add_combination():
        selected = [display_names[i] for i in listbox.curselection()]
        if len(selected) >= 2:
            combo_name = " + ".join(selected)
            combinations_listbox.insert(END, combo_name)
            selected_combinations.append(selected)
        else:
            messagebox.showwarning("Advertencia", "Selecciona al menos 2 productos para una combinación")
    
    def remove_combination():
        selected = combinations_listbox.curselection()
        if selected:
            combinations_listbox.delete(selected[0])
            selected_combinations.pop(selected[0])
    
    def finish_selection():
        root.quit()
        root.destroy()
    
    root = tk.Tk()
    root.title("Seleccionar Combinaciones de Productos")
    
    # Frame para selección de productos
    selection_frame = Frame(root)
    selection_frame.pack(pady=10, padx=10, fill=tk.X)
    
    Label(selection_frame, text="Productos disponibles:").pack()
    listbox = Listbox(selection_frame, selectmode=MULTIPLE, height=10, width=50)
    listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
    scrollbar = tk.Scrollbar(selection_frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    listbox.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=listbox.yview)
    
    for product in display_names:
        listbox.insert(END, product)
    
    # Frame para combinaciones seleccionadas
    combo_frame = Frame(root)
    combo_frame.pack(pady=10, padx=10, fill=tk.X)
    
    Label(combo_frame, text="Combinaciones seleccionadas:").pack()
    combinations_listbox = Listbox(combo_frame, height=5, width=50)
    combinations_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
    # Botones de acción
    button_frame = Frame(root)
    button_frame.pack(pady=10)
    
    Button(button_frame, text="Añadir Combinación", command=add_combination).pack(side=tk.LEFT, padx=5)
    Button(button_frame, text="Eliminar Combinación", command=remove_combination).pack(side=tk.LEFT, padx=5)
    Button(button_frame, text="Finalizar Selección", command=finish_selection).pack(side=tk.LEFT, padx=5)
    
    root.mainloop()
    return selected_combinations

def process_individual_products(df):
    """Procesa el reporte para productos individuales"""
    # Filtrar solo los productos que están en nuestro diccionario
    df = df[df['itemId'].isin(product_dict.keys())]
    
    # Mapear los itemId a nombres de productos
    df['product_name'] = df['itemId'].map(product_dict)
    
    # 1. Calcular usuarios que recompraron (compraron el mismo producto en diferentes órdenes)
    reordered_users = df.groupby(['customer_id', 'product_name'])['id'].nunique()
    reordered_users = reordered_users[reordered_users > 1].reset_index()
    reordered_users_count = reordered_users['product_name'].value_counts().reset_index()
    reordered_users_count.columns = ['product_name', 'usuarios_recompraron']
    
    # 2. Calcular total de usuarios que compraron cada producto
    total_users = df.groupby('product_name')['customer_id'].nunique().reset_index()
    total_users.columns = ['product_name', 'total_usuarios_compraron']
    
    # 3. Calcular número de órdenes distintas que contienen cada producto
    orders_with_product = df.groupby('product_name')['id'].nunique().reset_index()
    orders_with_product.columns = ['product_name', 'ordenes_con_producto']
    
    # Combinar todos los datos
    result = pd.merge(total_users, reordered_users_count, on='product_name', how='left')
    result = pd.merge(result, orders_with_product, on='product_name', how='left')
    
    # Rellenar NaN con 0 (para productos que no tengan recompras)
    result['usuarios_recompraron'] = result['usuarios_recompraron'].fillna(0).astype(int)
    
    # 4. Calcular porcentaje de recompra
    result['porcentaje_recompra'] = (result['usuarios_recompraron'] / result['total_usuarios_compraron'] * 100).round(2)
    
    # Ordenar por nombre de producto ascendente
    result = result.sort_values('product_name', ascending=True)
    
    # Reordenar columnas
    result = result[['product_name', 'usuarios_recompraron', 'total_usuarios_compraron', 
                    'ordenes_con_producto', 'porcentaje_recompra']]
    
    return result

def process_product_combinations(df, combinations):
    """Procesa el reporte para combinaciones de productos"""
    # Filtrar solo los productos que están en nuestro diccionario
    df = df[df['itemId'].isin(product_dict.keys())]
    
    # Mapear los itemId a nombres de productos
    df['product_name'] = df['itemId'].map(product_dict)
    
    # Preparar los resultados
    results = []
    
    for combo in combinations:
        combo_name = " + ".join(sorted(combo))  # Ordenar alfabéticamente los productos en la combinación
        
        # Filtrar órdenes que contienen TODOS los productos de la combinación
        orders_with_combo = df.groupby('id').filter(lambda x: all(p in x['product_name'].values for p in combo))
        
        if not orders_with_combo.empty:
            # 1. Calcular usuarios que recompraron la combinación (en diferentes órdenes)
            reordered_users = orders_with_combo.groupby('customer_id')['id'].nunique()
            reordered_users = reordered_users[reordered_users > 1].count()
            
            # 2. Calcular total de usuarios que compraron la combinación
            total_users = orders_with_combo['customer_id'].nunique()
            
            # 3. Calcular número de órdenes distintas que contienen la combinación
            total_orders = orders_with_combo['id'].nunique()
            
            # 4. Calcular porcentaje de recompra
            rebuy_percentage = (reordered_users / total_users * 100).round(2) if total_users > 0 else 0
            
            results.append({
                'product_name': combo_name,
                'usuarios_recompraron': reordered_users,
                'total_usuarios_compraron': total_users,
                'ordenes_con_producto': total_orders,
                'porcentaje_recompra': rebuy_percentage
            })
    
    # Crear DataFrame con los resultados
    result = pd.DataFrame(results)
    
    # Ordenar por nombre de combinación ascendente
    if not result.empty:
        result = result.sort_values('product_name', ascending=True)
        # Reordenar columnas para que coincidan con el otro reporte
        result = result[['product_name', 'usuarios_recompraron', 'total_usuarios_compraron', 
                        'ordenes_con_producto', 'porcentaje_recompra']]
    
    return result

def adjust_column_widths(writer, sheet_name, df):
    """Ajusta automáticamente el ancho de las columnas en el Excel"""
    worksheet = writer.sheets[sheet_name]
    
    for i, col in enumerate(df.columns):
        # Encontrar la longitud máxima del contenido de la columna
        max_len = max((
            df[col].astype(str).map(len).max(),  # Longitud máxima de los datos
            len(str(col))  # Longitud del nombre de la columna
        )) + 2  # Pequeño margen
        
        # Ajustar el ancho de la columna
        worksheet.set_column(i, i, max_len)

# [Todo el código anterior permanece igual hasta la función main()]

def main():
    # 1. Obtener fechas del selector
    start_date, end_date = open_rebuy_date_selector()
    
    if not start_date or not end_date:
        return
    
    # 2. Seleccionar tipo de reporte
    report_types = select_report_type()
    
    if not report_types:
        return
    
    # 3. Si se seleccionó combinaciones, obtener las combinaciones
    combinations = []
    if 'combinations' in report_types:
        combinations = select_product_combinations()
        if not combinations:
            messagebox.showwarning("Advertencia", "No se seleccionaron combinaciones. No se generará reporte de combinaciones.")
            report_types.remove('combinations')
            if not report_types:
                return
    
    # 4. Obtener datos de la primera consulta (todas las compras)
    query_all_orders = """
    WITH ordenes_con_beard AS (
        SELECT DISTINCT fo.id
        FROM bi.fact_orders fo
        JOIN bi.fact_sales_order_items oi ON fo.id = oi.salesOrderId
        WHERE oi.category = 'IG00000000000000000000000000000029'
    )

    SELECT fo.customer_id, fo.id, oi.itemId 
    FROM bi.fact_orders fo
    JOIN bi.fact_sales_order_items oi ON fo.id = oi.salesOrderId
    WHERE fo.status != 'CANCELLED'
    AND fo.order_plan = 'OTO'
    AND fo.recurrent = 0
    AND fo.id IN (SELECT id FROM ordenes_con_beard)
    """
    main_df = execute_query(query_all_orders)
    
    # 5. Obtener datos de la segunda consulta (primera compra en rango de fechas)
    query_first_orders = f"""
    SELECT fo.customer_id
    FROM bi.fact_orders fo
    JOIN bi.fact_sales_order_items oi ON fo.id = oi.salesOrderId
    WHERE fo.status != 'CANCELLED'
    AND fo.order_plan = 'OTO'
    AND fo.recurrent = 0
    AND fo.is_first_order = 1
    AND fo.created_at BETWEEN '{start_date}' AND '{end_date}'
    """
    first_time_buyers_df = execute_query(query_first_orders)
    
    if not main_df.empty and not first_time_buyers_df.empty:
        # Filtrar solo usuarios que están en el listado de primera compra
        valid_users = first_time_buyers_df['customer_id'].unique()
        filtered_df = main_df[main_df['customer_id'].isin(valid_users)]

        # Calcular el total de usuarios únicos que cumplen todos los criterios
        total_usuarios_muestra = filtered_df['customer_id'].nunique()
        print(f'El total de usuarios es: {total_usuarios_muestra}')
        
        # Procesar los reportes seleccionados
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import numbers
        
        wb = Workbook()
        
        if 'total' in report_types:
            individual_result = process_individual_products(filtered_df)
            ws1 = wb.create_sheet(title="Productos_Individuales")
            
            # Escribir los datos
            for r in dataframe_to_rows(individual_result, index=False, header=True):
                ws1.append(r)
            
            # Aplicar formato de porcentaje a la última columna (columna E)
            for row in ws1.iter_rows(min_row=2, min_col=5, max_col=5):
                for cell in row:
                    cell.number_format = '0.00%'
                    cell.value = cell.value / 100  # Convertir a decimal para formato %
            
            # Congelar paneles
            ws1.freeze_panes = 'A2'
            
            # Ajustar anchos de columna
            for column in ws1.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws1.column_dimensions[column_letter].width = adjusted_width
        
        if 'combinations' in report_types and combinations:
            combo_result = process_product_combinations(filtered_df, combinations)
            ws2 = wb.create_sheet(title="Combinaciones")
            
            # Escribir los datos
            for r in dataframe_to_rows(combo_result, index=False, header=True):
                ws2.append(r)
            
            # Aplicar formato de porcentaje a la última columna (columna E)
            for row in ws2.iter_rows(min_row=2, min_col=5, max_col=5):
                for cell in row:
                    cell.number_format = '0.00%'
                    cell.value = cell.value / 100  # Convertir a decimal para formato %
            
            # Congelar paneles
            ws2.freeze_panes = 'A2'
            
            # Ajustar anchos de columna
            for column in ws2.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws2.column_dimensions[column_letter].width = adjusted_width
        
        # Eliminar la hoja por defecto si no se usa
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Guardar el archivo
        wb.save(f'reporte_recompras_usuarios_{start_date}_a_{end_date}.xlsx')
        print("Reporte generado exitosamente")
    else:
        print("No se encontraron datos para generar el reporte")

if __name__ == "__main__":
    main()